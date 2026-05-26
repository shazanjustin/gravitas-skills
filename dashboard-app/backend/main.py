from fastapi import FastAPI, Query, HTTPException, Response
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from typing import Optional
from data_processor import load_data
import pandas as pd
import math
import json
import os
import io
import requests as http_requests
from dotenv import load_dotenv

load_dotenv()  # Load GROQ_API_KEY from .env

app = FastAPI(title="Social Media Analytics API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], 
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

CACHE = {
    "data": None,
    "last_fetched": None
}

def get_filtered_data(start_date: Optional[str] = None, end_date: Optional[str] = None):
    if CACHE["data"] is None:
        raise HTTPException(status_code=400, detail="Data not loaded. Please provide a sheet URL to load data.")
    
    df = pd.DataFrame(CACHE["data"])
    df['date'] = pd.to_datetime(df['date'])
    
    if start_date:
        df = df[df['date'] >= pd.to_datetime(start_date)]
    if end_date:
        end_dt = pd.to_datetime(end_date)
        if end_dt.time() == pd.Timestamp('00:00:00').time():
            end_dt = end_dt + pd.Timedelta(days=1, seconds=-1)
        df = df[df['date'] <= end_dt]
        
    return df

@app.get("/api/refresh")
def refresh_data(sheet_url: str = Query(...)):
    try:
        CACHE["data"] = load_data(sheet_url)
        CACHE["last_fetched"] = pd.Timestamp.now().isoformat()
        return {"message": "Data refreshed successfully", "last_fetched": CACHE["last_fetched"]}
    except Exception as e:
        return {"error": f"Failed to load data from sheet: {str(e)}"}

@app.get("/api/dashboard-summary")
def get_dashboard_summary(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None)):
    df = get_filtered_data(start_date, end_date)
    
    total_reach = df['reach'].sum()
    total_engagement = df['engagement'].sum()
    avg_engagement_rate = (total_engagement / total_reach * 100) if total_reach > 0 else 0
    
    platform_group = df.groupby('platform').agg({'engagement': 'sum'}).reset_index()
    top_platform = platform_group.sort_values(by='engagement', ascending=False).iloc[0]['platform'] if not platform_group.empty else "N/A"
    
    return {
        "kpis": {
            "total_reach": float(total_reach),
            "total_engagement": float(total_engagement),
            "avg_engagement_rate": float(avg_engagement_rate),
            "top_platform": str(top_platform)
        }
    }

@app.get("/api/platform-stats")
def get_platform_stats(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None)):
    df = get_filtered_data(start_date, end_date)
    
    stats = {}
    for platform in ['Facebook', 'Instagram', 'TikTok', 'YouTube', 'LinkedIn']:
        pdf = df[df['platform'] == platform]
        if pdf.empty:
            stats[platform] = None
            continue
        
        posts_count = len(pdf)
        avg_reach = float(pdf['reach'].mean()) if posts_count > 0 else 0
        avg_er = float(pdf['engagement_rate'].mean()) if posts_count > 0 else 0
            
        stats[platform] = {
            "posts_count": posts_count,
            "avg_reach": avg_reach,
            "avg_engagement_rate": avg_er
        }
        
    return stats

@app.get("/api/organic-content")
def get_organic_content(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None)):
    df = get_filtered_data(start_date, end_date)
    
    def clean_records(df_subset):
        res = []
        for r in df_subset.to_dict(orient='records'):
            clean_r = {k: (None if (isinstance(v, float) and math.isnan(v)) else v) for k,v in r.items()}
            if pd.notnull(clean_r['date']):
                clean_r['date'] = str(clean_r['date'])
            res.append(clean_r)
        return res

    result = {}
    for platform in ['Facebook', 'Instagram', 'TikTok', 'YouTube', 'LinkedIn']:
        org = df[(df['platform'] == platform) & (df['is_organic'] == True)]
        
        if org.empty:
            result[platform] = {"top": [], "bottom": []}
            continue
            
        org = org.sort_values(by='engagement', ascending=False)
        
        top5 = clean_records(org.head(5))
        bottom5 = clean_records(org.tail(5))
        
        result[platform] = {
            "top": top5,
            "bottom": bottom5
        }
    
    return result


@app.get("/api/engagement-summary")
def get_engagement_summary(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None)):
    df = get_filtered_data(start_date, end_date)

    platforms = ['Facebook', 'Instagram', 'TikTok', 'YouTube', 'LinkedIn']
    platform_data = []
    overall_total = 0
    overall_posts = 0

    for platform in platforms:
        pdf = df[df['platform'] == platform]
        if pdf.empty:
            continue
        total_eng = float(pdf['engagement'].sum())
        posts = len(pdf)
        avg_eng = total_eng / posts if posts > 0 else 0
        overall_total += total_eng
        overall_posts += posts
        platform_data.append({
            "platform": platform,
            "total_engagement": total_eng,
            "posts_count": posts,
            "avg_engagement_per_post": avg_eng,
        })

    return {
        "platforms": platform_data,
        "overall": {
            "total_engagement": overall_total,
            "posts_count": overall_posts,
            "avg_engagement_per_post": overall_total / overall_posts if overall_posts > 0 else 0,
        }
    }

# ---------------------------------------------------------------------------
# Content-type classification
# ---------------------------------------------------------------------------
# Ordered priority list: first matching rule wins.
CONTENT_TYPE_RULES = [
    ("Raya / Festive",         ["raya", "hari raya", "aidilfitri", "aidiladha", "ramadan", "festive", "cny", "chinese new year", "deepavali", "diwali", "christmas", "new year", "merdeka", "hari kebangsaan"]),
    ("Creator Collaboration",  ["collab", "collaboration", "creator", "influencer", "x ", "feat.", "featuring", "bersama", "ft."]),
    ("Security / Fraud Alert", ["scam", "fraud", "phishing", "jangan mudah", "protect", "security", "selamat", "secure", "alert", "penipuan", "beware", "warning"]),
    ("App Features",           ["app", "cimb clicks", "mobile banking", "digital banking", "online banking", "feature", "fitur", "clicks app"]),
    ("App Tutorial",           ["tutorial", "how to", "how-to", "cara ", "step by step", "panduan", "guide", "langkah", "learn how"]),
    ("Financial Literacy",     ["financial literacy", "kewangan", "money tip", "tip kewangan", "budgeting", "bajet", "savings", "simpanan", "investment", "pelaburan", "financial planning", "wealth", "unit trust", "amanah saham", "finance tip", "financial tip", "did you know", "tahukah anda"]),
    ("Product Promotion",      ["promo", "promotion", "offer", "deal", "discount", "cashback", "reward", "kredit", "credit card", "kad kredit", "loan", "pinjaman", "mortgage", "home loan", "personal loan", "rate", "kadar", "apply now", "daftar sekarang", "special rate"]),
    ("Market Outlook",         ["market outlook", "economy", "economic", "gdp", "inflation", "interest rate", "bnm", "bank negara", "quarter", "q1", "q2", "q3", "q4", "forecast", "ringgit", "bursa", "klse"]),
    ("Awards / Recognition",   ["award", "recognition", "winner", "excellence", "best bank", "achievement", "accolade", "rated", "ranking", "ranked"]),
    ("Announcements",          ["announcement", "announce", "introducing", "new launch", "launching", "press release", "media release"]),
    ("Brand / CSR",            ["csr", "community", "sustainability", "environment", "charity", "donation", "green", "social responsibility", "gotong royong", "kesukarelawan", "volunteer", "wakaf", "zakat", "sedekah"]),
    ("ASEAN Culture",          ["asean", "malaysia", "malaysian", "budaya", "heritage", "tradition", "warisan", "kebudayaan", "local culture"]),
    ("Event Recap",            ["event", "recap", "highlight", "ceremony", "launch event", "conference", "summit", "forum", "seminar", "webinar", "workshop"]),
    ("Interactive",            ["quiz", "poll", "challenge", "contest", "giveaway", "win ", "tag a friend", "share your", "comment below", "tell us", "vote"]),
    ("Talent / Career",        ["career", "hiring", "job", "talent", "employee", "kerjaya", "team member", "join us", "we are hiring", "internship"]),
    ("SME / Business Banking", ["sme", "business", "entrepreneur", "enterprise", "corporate", "trade finance", "usahawan", "perniagaan", "b2b"]),
    ("Campaign Video",         ["campaign", "video campaign", "tv ad", "advertisement", "ad "]),
]
FALLBACK_TYPE = "General / Other"


def classify_content_type(title: str) -> str:
    """Return the first matching content-type label for a post title."""
    if not title or title.strip().lower() in ("", "nan", "none"):
        return FALLBACK_TYPE
    text = title.lower()
    for label, keywords in CONTENT_TYPE_RULES:
        if any(kw in text for kw in keywords):
            return label
    return FALLBACK_TYPE


@app.get("/api/content-types")
def get_content_types(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None)):
    """Return content-type breakdown (post counts) per platform + overall."""
    df = get_filtered_data(start_date, end_date)

    platforms = ['Facebook', 'Instagram', 'TikTok', 'YouTube', 'LinkedIn']
    result = {}
    overall_counts: dict[str, int] = {}

    for platform in platforms:
        pdf = df[df['platform'] == platform]
        if pdf.empty:
            result[platform] = []
            continue

        counts: dict[str, int] = {}
        for title in pdf['title'].fillna(''):
            ct = classify_content_type(str(title))
            counts[ct] = counts.get(ct, 0) + 1

        # Sort by count desc
        sorted_counts = sorted(counts.items(), key=lambda x: x[1], reverse=True)
        result[platform] = [{"type": t, "count": c} for t, c in sorted_counts]

        for t, c in counts.items():
            overall_counts[t] = overall_counts.get(t, 0) + c

    result["Overall"] = [
        {"type": t, "count": c}
        for t, c in sorted(overall_counts.items(), key=lambda x: x[1], reverse=True)
    ]
    return result


@app.get("/api/all-content")
def get_all_content(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None)):
    df = get_filtered_data(start_date, end_date)

    def clean_records(df_subset):
        res = []
        for r in df_subset.to_dict(orient='records'):
            clean_r = {k: (None if (isinstance(v, float) and math.isnan(v)) else v) for k,v in r.items()}
            if pd.notnull(clean_r['date']):
                clean_r['date'] = str(clean_r['date'])
            res.append(clean_r)
        return res

    result = {}
    for platform in ['Facebook', 'Instagram', 'TikTok', 'YouTube', 'LinkedIn']:
        org = df[(df['platform'] == platform) & (df['is_organic'] == True)]
        if org.empty:
            result[platform] = []
            continue
        org = org.sort_values(by='engagement', ascending=False)
        result[platform] = clean_records(org)

    return result


@app.get("/api/format-performance")
def get_format_performance(start_date: Optional[str] = Query(None), end_date: Optional[str] = Query(None)):
    df = get_filtered_data(start_date, end_date)
    
    platforms = ['Facebook', 'Instagram', 'LinkedIn', 'TikTok', 'YouTube']
    result = []
    
    grand_posts = 0
    grand_reach = 0
    grand_eng = 0
    grand_er_sum = 0
    
    for platform in platforms:
        pdf = df[(df['platform'] == platform) & (df['is_organic'] == True)]
        if pdf.empty:
            continue
            
        platform_posts = 0
        platform_reach = 0
        platform_eng = 0
        platform_er_sum = 0
        
        formats = pdf['format'].unique()
        for fmt in sorted(formats):
            fdf = pdf[pdf['format'] == fmt]
            posts = len(fdf)
            if posts == 0: continue
            
            reach_sum = fdf['reach'].sum()
            eng_sum = fdf['engagement'].sum()
            er_mean = fdf['engagement_rate'].mean()
            
            platform_posts += posts
            platform_reach += reach_sum
            platform_eng += eng_sum
            platform_er_sum += er_mean * posts
            
            result.append({
                "platform": platform,
                "format": fmt,
                "is_total": False,
                "posts": posts,
                "avg_reach": reach_sum / posts,
                "avg_engagement": eng_sum / posts,
                "avg_er": er_mean
            })
            
        if platform_posts > 0:
            result.append({
                "platform": platform,
                "format": f"{platform} Total",
                "is_total": True,
                "posts": platform_posts,
                "avg_reach": platform_reach / platform_posts,
                "avg_engagement": platform_eng / platform_posts,
                "avg_er": platform_er_sum / platform_posts
            })
            
            grand_posts += platform_posts
            grand_reach += platform_reach
            grand_eng += platform_eng
            grand_er_sum += platform_er_sum
            
    if grand_posts > 0:
        result.append({
            "platform": "Grand Total",
            "format": "",
            "is_total": True,
            "is_grand_total": True,
            "posts": grand_posts,
            "avg_reach": grand_reach / grand_posts,
            "avg_engagement": grand_eng / grand_posts,
            "avg_er": grand_er_sum / grand_posts
        })
        
    return result


@app.get("/api/export-all-contents")
def export_all_contents(
    sheet_url: str = Query(..., description="Public Google Sheet URL"),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None)
):
    """Return an Excel file with all contents (organic + paid) for the given date range."""
    import re
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', sheet_url)
    if not match:
        return {"error": "Invalid Google Sheet URL. Could not extract spreadsheet ID."}
    sheet_id = match.group(1)
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"

    try:
        xl = pd.ExcelFile(export_url)
    except Exception as e:
        return {"error": f"Failed to download Google Sheet: {str(e)}"}

    def _get(row, *keys, default=''):
        for k in keys:
            v = row.get(k)
            if v is not None and not (isinstance(v, float) and math.isnan(v)):
                return v
        return default

    def _num(row, *keys):
        for k in keys:
            v = row.get(k)
            try:
                f = float(v)
                if not math.isnan(f): return f
            except: pass
        return 0.0

    def _in_range(dt):
        if pd.isna(dt): return False
        if start_date and dt < pd.to_datetime(start_date): return False
        if end_date:
            end_dt = pd.to_datetime(end_date) + pd.Timedelta(days=1, seconds=-1)
            if dt > end_dt: return False
        return True

    rows = []

    # ── Facebook ────────────────────────────────────────────────────────────────
    if 'Raw_FB' in xl.sheet_names:
        fb = xl.parse('Raw_FB')
        for _, row in fb.iterrows():
            dt = pd.to_datetime(row.get('Publish time'), errors='coerce')
            if not _in_range(dt): continue
            views = _num(row, 'Views')
            reach = _num(row, 'Reach', 'Lifetime Post Total Reach')
            reactions = _num(row, 'Reactions, comments and shares')
            comments = _num(row, 'Comments')
            shares = _num(row, 'Shares')
            likes = max(0, reactions - comments - shares)
            total_eng = reactions
            er = round(total_eng / views * 100, 2) if views > 0 else 0
            rows.append({
                'Date(Publish)': dt.date(),
                'Platform': 'Facebook',
                'Format': str(_get(row, 'Post type')),
                'Pillar': '',
                'Organic/Paid': str(_get(row, 'Organic/Paid')),
                'Collab': str(_get(row, 'Collab')),
                'Title': '',
                'Caption': str(_get(row, 'Description', 'Post message', 'Title')),
                'Reach': reach,
                'Views': views,
                'Interaction': total_eng,
                'ER%': er,
                'Likes': likes,
                'Comments': comments,
                'Shares': shares,
                'Saves': '',
                'Reposts': '',
                'URL': str(_get(row, 'Permalink', 'Link')),
                'Year Month': dt.strftime('%Y %B'),
            })

    # ── Instagram ───────────────────────────────────────────────────────────────
    if 'Raw_IG' in xl.sheet_names:
        ig = xl.parse('Raw_IG')
        for _, row in ig.iterrows():
            dt = pd.to_datetime(row.get('Publish time'), errors='coerce')
            if not _in_range(dt): continue
            views = _num(row, 'Views')
            reach = _num(row, 'Reach')
            likes = _num(row, 'Likes')
            comments = _num(row, 'Comments')
            shares = _num(row, 'Shares')
            saves = _num(row, 'Saves')
            total_eng = likes + comments + shares + saves
            er = round(total_eng / views * 100, 2) if views > 0 else 0
            rows.append({
                'Date(Publish)': dt.date(),
                'Platform': 'Instagram',
                'Format': str(_get(row, 'Post type')),
                'Pillar': '',
                'Organic/Paid': str(_get(row, 'Organic/Paid')),
                'Collab': str(_get(row, 'Collab')),
                'Title': '',
                'Caption': str(_get(row, 'Description')),
                'Reach': reach,
                'Views': views,
                'Interaction': total_eng,
                'ER%': er,
                'Likes': likes,
                'Comments': comments,
                'Shares': shares,
                'Saves': saves,
                'Reposts': '',
                'URL': str(_get(row, 'Permalink')),
                'Year Month': dt.strftime('%Y %B'),
            })

    # ── YouTube ─────────────────────────────────────────────────────────────────
    if 'Raw_Youtube' in xl.sheet_names:
        yt = xl.parse('Raw_Youtube')
        for _, row in yt.iterrows():
            dt = pd.to_datetime(row.get('Video publish time'), errors='coerce')
            if not _in_range(dt): continue
            views = _num(row, 'Views')
            likes = _num(row, 'Likes')
            comments = _num(row, 'Comments added')
            shares = _num(row, 'Shares')
            total_eng = likes + comments + shares
            er = round(total_eng / views * 100, 2) if views > 0 else 0
            rows.append({
                'Date(Publish)': dt.date(),
                'Platform': 'YouTube',
                'Format': 'Video',
                'Pillar': '',
                'Organic/Paid': str(_get(row, 'Organic/ Paid', 'Organic/Paid')),
                'Collab': str(_get(row, 'Collab')),
                'Title': '',
                'Caption': str(_get(row, 'Video title')),
                'Reach': views,
                'Views': views,
                'Interaction': total_eng,
                'ER%': er,
                'Likes': likes,
                'Comments': comments,
                'Shares': shares,
                'Saves': '',
                'Reposts': '',
                'URL': f"https://youtube.com/watch?v={_get(row, 'Content')}",
                'Year Month': dt.strftime('%Y %B'),
            })

    # ── TikTok ──────────────────────────────────────────────────────────────────
    if 'Raw_Tiktok' in xl.sheet_names:
        tt = xl.parse('Raw_Tiktok')
        for _, row in tt.iterrows():
            dt = pd.to_datetime(row.get('Post time'), errors='coerce')
            if not _in_range(dt): continue
            views = _num(row, 'Video views')
            if views < 10: continue   # keep same filter as main pipeline
            likes = _num(row, 'Likes')
            comments = _num(row, 'Comments')
            shares = _num(row, 'Shares')
            favorites = _num(row, 'Add to Favorites')
            total_eng = likes + comments + shares + favorites
            er = round(total_eng / views * 100, 2) if views > 0 else 0
            rows.append({
                'Date(Publish)': dt.date(),
                'Platform': 'TikTok',
                'Format': 'Video',
                'Pillar': '',
                'Organic/Paid': str(_get(row, 'Organic/Paid')),
                'Collab': str(_get(row, 'Collab')),
                'Title': '',
                'Caption': str(_get(row, 'Video description', 'Video title')),
                'Reach': views,
                'Views': views,
                'Interaction': total_eng,
                'ER%': er,
                'Likes': likes,
                'Comments': comments,
                'Shares': shares,
                'Saves': favorites,
                'Reposts': '',
                'URL': str(_get(row, 'Video link')),
                'Year Month': dt.strftime('%Y %B'),
            })

    # ── LinkedIn ────────────────────────────────────────────────────────────────
    if 'Raw_LI' in xl.sheet_names:
        li = xl.parse('Raw_LI')
        for _, row in li.iterrows():
            dt = pd.to_datetime(row.get('Created date'), errors='coerce')
            if not _in_range(dt): continue
            impressions = _num(row, 'Impressions')
            views = _num(row, 'Views')
            likes = _num(row, 'Likes')
            comments = _num(row, 'Comments')
            reposts = _num(row, 'Reposts')
            total_eng = likes + comments + reposts
            er = round(total_eng / impressions * 100, 2) if impressions > 0 else 0
            rows.append({
                'Date(Publish)': dt.date(),
                'Platform': 'LinkedIn',
                'Format': str(_get(row, 'Content Type')),
                'Pillar': '',
                'Organic/Paid': str(_get(row, 'Organic/Paid')),
                'Collab': str(_get(row, 'Collab')),
                'Title': '',
                'Caption': str(_get(row, 'Post content', 'Update title', 'Post title')),
                'Reach': impressions,
                'Views': views,
                'Interaction': total_eng,
                'ER%': er,
                'Likes': likes,
                'Comments': comments,
                'Shares': '',
                'Saves': '',
                'Reposts': reposts,
                'URL': str(_get(row, 'Post link')),
                'Year Month': dt.strftime('%Y %B'),
            })

    # ── Build DataFrame & sort by date ──────────────────────────────────────────
    df_export = pd.DataFrame(rows, columns=[
        'Date(Publish)', 'Platform', 'Format', 'Pillar', 'Organic/Paid',
        'Collab', 'Title', 'Caption', 'Reach', 'Views', 'Interaction',
        'ER%', 'Likes', 'Comments', 'Shares', 'Saves', 'Reposts', 'URL', 'Year Month'
    ])
    df_export = df_export.sort_values('Date(Publish)', ascending=True).reset_index(drop=True)

    # ── Write to Excel in memory ─────────────────────────────────────────────────
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='All Contents')
        ws = writer.sheets['All Contents']
        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    output.seek(0)

    period = f"{start_date or 'all'}_{end_date or 'present'}"
    filename = f"CIMB_All_Contents_{period}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@app.get("/api/pillar-er")
def get_pillar_er(
    sheet_url: str = Query(..., description="Public Google Sheet URL"),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None)
):
    """
    Read the 'All Contents' sheet from the user-supplied Google Sheet URL,
    filter to Organic rows only, and return avg ER% pivot: Pillar × Platform.
    """
    # Convert any edit/view URL to an xlsx export URL
    import re
    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', sheet_url)
    if not match:
        return {"error": "Invalid Google Sheet URL. Could not extract spreadsheet ID."}
    sheet_id = match.group(1)
    export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"

    try:
        xl = pd.ExcelFile(export_url)
    except Exception as e:
        return {"error": f"Could not read the Google Sheet. Make sure it is publicly shared. ({e})"}

    # Find the correct sheet — try common names
    target_sheet = None
    for candidate in ["All Contents", "All Content", "All contents", "all contents"]:
        if candidate in xl.sheet_names:
            target_sheet = candidate
            break
    if target_sheet is None:
        return {
            "error": f"Could not find an 'All Contents' sheet. Sheets found: {xl.sheet_names}"
        }

    df = xl.parse(target_sheet)

    # Normalise column names (strip whitespace)
    df.columns = [str(c).strip() for c in df.columns]

    required = {"Pillar", "Platform", "Organic/Paid", "ER%"}
    missing = required - set(df.columns)
    if missing:
        # Try alternative ER column names
        for alt in ["ER %", "Engagement Rate", "Engagement Rate %"]:
            if alt in df.columns:
                df.rename(columns={alt: "ER%"}, inplace=True)
                missing = required - set(df.columns)
                break
    if missing:
        return {"error": f"Required columns missing from sheet: {missing}. Found: {df.columns.tolist()}"}

    # Date filter
    if "Date(Publish)" in df.columns:
        df["Date(Publish)"] = pd.to_datetime(df["Date(Publish)"], errors="coerce")
        if start_date:
            df = df[df["Date(Publish)"] >= pd.to_datetime(start_date)]
        if end_date:
            end_dt = pd.to_datetime(end_date) + pd.Timedelta(days=1, seconds=-1)
            df = df[df["Date(Publish)"] <= end_dt]

    # Filter organic only
    organic_mask = df["Organic/Paid"].astype(str).str.strip().str.lower() == "organic"
    df = df[organic_mask].copy()

    # Clean ER% — strip % signs, coerce to float
    df["ER%"] = (
        df["ER%"].astype(str)
        .str.replace("%", "", regex=False)
        .str.strip()
        .pipe(pd.to_numeric, errors="coerce")
    )

    # Drop rows with no Pillar, no Platform, or no ER
    df = df.dropna(subset=["Pillar", "Platform", "ER%"])
    df = df[df["Pillar"].astype(str).str.strip() != ""]

    # Build pivot: rows = Pillar, cols = Platform, values = mean ER%
    platforms = ["Facebook", "Instagram", "LinkedIn", "TikTok", "YouTube"]
    pillars = sorted(df["Pillar"].astype(str).str.strip().unique())

    pivot = []
    for pillar in pillars:
        pdata = {"pillar": pillar}
        pf = df[df["Pillar"].astype(str).str.strip() == pillar]
        for plat in platforms:
            plat_df = pf[pf["Platform"].astype(str).str.strip() == plat]
            if plat_df.empty:
                pdata[plat] = None
            else:
                pdata[plat] = round(float(plat_df["ER%"].mean()), 2)
        pivot.append(pdata)

    return {"platforms": platforms, "pivot": pivot}


@app.get("/api/export-cross-platform")
def export_cross_platform(
    sheet_url: str = Query(...),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None)
):
    import re
    from openpyxl.styles import PatternFill, Font, Alignment

    match = re.search(r'/spreadsheets/d/([a-zA-Z0-9_-]+)', sheet_url)
    if not match:
        return {"error": "Invalid Google Sheet URL."}
    export_url = f"https://docs.google.com/spreadsheets/d/{match.group(1)}/export?format=xlsx"

    try:
        xl = pd.ExcelFile(export_url)
    except Exception as e:
        return {"error": f"Could not read sheet: {e}"}

    target = next((s for s in xl.sheet_names if s.lower().startswith("all content")), None)
    if not target:
        return {"error": f"No 'All Contents' sheet found. Sheets: {xl.sheet_names}"}

    df = xl.parse(target)
    df.columns = [str(c).strip() for c in df.columns]

    # Date filter
    if "Date(Publish)" in df.columns:
        df["Date(Publish)"] = pd.to_datetime(df["Date(Publish)"], errors="coerce")
        if start_date:
            df = df[df["Date(Publish)"] >= pd.to_datetime(start_date)]
        if end_date:
            end_dt = pd.to_datetime(end_date) + pd.Timedelta(days=1, seconds=-1)
            df = df[df["Date(Publish)"] <= end_dt]

    # Normalise ER%
    if "ER%" in df.columns:
        df["ER%"] = pd.to_numeric(
            df["ER%"].astype(str).str.replace("%", "", regex=False).str.strip(),
            errors="coerce"
        )

    PLATFORMS = ["Facebook", "Instagram", "LinkedIn", "TikTok", "YouTube"]
    URL_COLS = {
        "Instagram": "IG URL", "TikTok": "TT URL",
        "Facebook": "FB URL", "YouTube": "YT URL", "LinkedIn": "LI URL"
    }

    # Sort by Pillar so same-pillar rows stay together
    sort_cols = ["Pillar"] + (["Date(Publish)"] if "Date(Publish)" in df.columns else [])
    df = df.sort_values(sort_cols, na_position="last").reset_index(drop=True)

    # Build one output row per unique normalized Caption
    seen = {}          # norm_cap -> first-seen order index
    groups = {}        # norm_cap -> {original_caption, pillar, title, plat_er, plat_url}

    for _, row in df.iterrows():
        caption = str(row.get("Caption", "")).strip()
        if not caption or caption == "nan":
            continue
            
        # Normalize the caption: ignore newlines, punctuation, whitespace
        norm_cap = re.sub(r'[\W_]+', '', caption).lower()
        if not norm_cap:
            norm_cap = caption.strip().lower()
            
        if norm_cap not in groups:
            seen[norm_cap] = len(seen)
            groups[norm_cap] = {
                "original_caption": caption,
                "pillar": str(row.get("Pillar", "") or "").strip(),
                "title":  str(row.get("Title", "")  or "").strip(),
                "er":  {},
                "url": {}
            }
        g = groups[norm_cap]
        plat = str(row.get("Platform", "")).strip()
        if plat in PLATFORMS and plat not in g["er"]:
            er_val = row.get("ER%")
            if er_val is not None and not (isinstance(er_val, float) and math.isnan(er_val)):
                g["er"][plat] = float(er_val)
            url_val = row.get("URL", row.get("URL.1", ""))
            g["url"][plat] = str(url_val) if url_val and str(url_val) != "nan" else ""

    # Build ordered output (sorted by pillar then first-seen)
    ordered = sorted(groups.keys(), key=lambda c: (groups[c]["pillar"], seen[c]))

    output_rows = []
    i = 0
    while i < len(ordered):
        pillar = groups[ordered[i]]["pillar"]
        # Collect all normalized captions for this pillar
        pillar_captions = []
        while i < len(ordered) and groups[ordered[i]]["pillar"] == pillar:
            pillar_captions.append(ordered[i])
            i += 1

        first_in_pillar = True
        # Accumulate platform ER values across this pillar for totals
        pillar_plat_ers = {p: [] for p in PLATFORMS}

        for norm_cap in pillar_captions:
            g = groups[norm_cap]
            valid_ers = [v for v in g["er"].values() if v is not None]
            grand_total = round(sum(valid_ers) / len(valid_ers), 2) if valid_ers else None

            output_rows.append({
                "Pillar":      pillar if first_in_pillar else "",
                "New Title":   g["title"] if g["title"] not in ("", "nan") else "",
                "Caption":     g["original_caption"],
                "Facebook":    g["er"].get("Facebook"),
                "Instagram":   g["er"].get("Instagram"),
                "LinkedIn":    g["er"].get("LinkedIn"),
                "TikTok":      g["er"].get("TikTok"),
                "YouTube":     g["er"].get("YouTube"),
                "Grand Total": grand_total,
                "IG URL":      g["url"].get("Instagram", ""),
                "TT URL":      g["url"].get("TikTok", ""),
                "FB URL":      g["url"].get("Facebook", ""),
                "YT URL":      g["url"].get("YouTube", ""),
                "LI URL":      g["url"].get("LinkedIn", ""),
                "_is_total":   False,
            })
            first_in_pillar = False

            for p in PLATFORMS:
                if g["er"].get(p) is not None:
                    pillar_plat_ers[p].append(g["er"][p])

        # Pillar Total row — average ER% per platform (2 dp)
        total_er = {
            p: round(sum(vals) / len(vals), 2) if vals else None
            for p, vals in pillar_plat_ers.items()
        }
        all_valid = [v for v in total_er.values() if v is not None]
        total_grand = round(sum(all_valid) / len(all_valid), 2) if all_valid else None

        output_rows.append({
            "Pillar":      f"{pillar} Total",
            "New Title":   "",
            "Caption":     "",
            "Facebook":    total_er.get("Facebook"),
            "Instagram":   total_er.get("Instagram"),
            "LinkedIn":    total_er.get("LinkedIn"),
            "TikTok":      total_er.get("TikTok"),
            "YouTube":     total_er.get("YouTube"),
            "Grand Total": total_grand,
            "IG URL":      "",
            "TT URL":      "",
            "FB URL":      "",
            "YT URL":      "",
            "LI URL":      "",
            "_is_total":   True,
        })

    COLS = ["Pillar","New Title","Caption","Facebook","Instagram","LinkedIn",
            "TikTok","YouTube","Grand Total","IG URL","TT URL","FB URL","YT URL","LI URL"]

    is_total_flags = [r.pop("_is_total") for r in output_rows]
    df_out = pd.DataFrame(output_rows, columns=COLS)

    # URL column indices (0-based in df = 1-based + 1 header in ws)
    URL_COL_NAMES = ["IG URL","TT URL","FB URL","YT URL","LI URL"]
    url_col_letters = {
        name: chr(ord('A') + COLS.index(name))
        for name in URL_COL_NAMES
    }

    # Write Excel
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="Cross Platform Contents Perform")
        ws = writer.sheets["Cross Platform Contents Perform"]

        from openpyxl.styles import Border, Side
        grey_fill  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        total_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        total_font  = Font(bold=True, color="000000")
        link_font   = Font(color="0563C1", underline="single")

        # Header row — light grey, bold
        for cell in ws[1]:
            cell.fill = grey_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Data rows
        for row_idx, is_total in enumerate(is_total_flags, start=2):
            if is_total:
                for cell in ws[row_idx]:
                    cell.fill = total_fill
                    cell.font = total_font

        # Make URL cells clickable hyperlinks
        for col_name, col_letter in url_col_letters.items():
            col_idx = COLS.index(col_name) + 1  # openpyxl is 1-based
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                url_val = str(cell.value or "").strip()
                if url_val and url_val.startswith("http"):
                    cell.hyperlink = url_val
                    cell.value = url_val
                    cell.font = link_font

        # Auto-width
        for col in ws.columns:
            width = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(width + 4, 80)

    buf.seek(0)
    period = f"{start_date or 'all'}_{end_date or 'present'}"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Cross_Platform_Contents_Performance_{period}.xlsx"'}
    )


@app.get("/api/executive-summary")
def get_executive_summary(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None)
):
    """Generate an AI executive summary for the selected date range using Gemini."""
    api_key = os.getenv("GROQ_API_KEY", "")
    if not api_key:
        return {"error": "GROQ_API_KEY environment variable is not set. Get a free key at https://console.groq.com"}

    df = get_filtered_data(start_date, end_date)
    if df.empty:
        return {"error": "No data available for the selected period."}

    # ── Collect per-platform analytics ─────────────────────────────────────────
    platforms = ['Facebook', 'Instagram', 'TikTok', 'YouTube', 'LinkedIn']
    platform_data = {}

    for platform in platforms:
        pdf = df[df['platform'] == platform]
        if pdf.empty:
            continue

        posts = len(pdf)
        total_eng = float(pdf['engagement'].sum())
        avg_reach = float(pdf['reach'].mean())
        avg_er = round(float(pdf['engagement_rate'].mean()), 2)

        # Top 3 posts
        top3 = pdf.sort_values('engagement', ascending=False).head(3)
        top_posts = [
            {"title": str(r['title'])[:80], "engagement": int(r['engagement']), "er": round(float(r['engagement_rate']), 2)}
            for _, r in top3.iterrows()
        ]

        # Content type distribution — exclude "General / Other"
        ct_counts: dict = {}
        for title in pdf['title'].fillna(''):
            ct = classify_content_type(str(title))
            if ct == FALLBACK_TYPE:
                continue  # skip General / Other
            ct_counts[ct] = ct_counts.get(ct, 0) + 1
        top_content_types = [t for t, _ in sorted(ct_counts.items(), key=lambda x: x[1], reverse=True)[:5]]

        # NOTE: top_posts deliberately excluded — no post names reach the AI
        platform_data[platform] = {
            "posts": posts,
            "total_engagement": int(total_eng),
            "avg_reach": int(avg_reach),
            "avg_er_pct": avg_er,
            "top_content_types": top_content_types,
        }

    # Pre-rank platforms by avg ER so AI gets explicit ordering
    er_ranking = sorted(
        [(p, d["avg_er_pct"]) for p, d in platform_data.items()],
        key=lambda x: x[1], reverse=True
    )
    rank_label = ", ".join([f"#{i+1} {p} ({er}% ER)" for i, (p, er) in enumerate(er_ranking)])

    period_label = f"{start_date or 'beginning'} to {end_date or 'present'}"

    # ── Build prompt ────────────────────────────────────────────────────────────
    prompt = f"""You are a senior social media strategist writing an executive summary for CIMB Bank Malaysia's management team.
The summary covers the period: {period_label}.

Platform ER Ranking (by Avg ER%): {rank_label}

Platform Performance Data:
{json.dumps(platform_data, indent=2)}

Return ONLY valid JSON (no markdown, no code blocks) with this exact structure:
{{
  "top_platform": "<name of #1 platform by Avg ER%>",
  "top_platform_reason": "<one concise sentence — state its Avg ER% and why it led>",
  "key_highlights": [
    "<POINT 1 — The #1 platform by Avg ER%. State: platform name in bold (**Name**), its Avg ER%, total engagement, and top content types that drove it. One clear sentence. No post names.>",
    "<POINT 2 — The #2 platform by Avg ER%. Same format: platform name bold, Avg ER%, total engagement, top content types. One clear sentence. No post names.>"
  ],
  "audience_behaviour": [
    "<Insight 1 — A broad, strategic observation about what type of content or format audiences responded to across platforms this period. No post names. Management-level language.>",
    "<Insight 2 — Another strategic behavioural trend, e.g. about format preferences (short-form vs long-form), content themes, or engagement patterns. No post names.>"
  ],
  "recommendations": {{
    "Facebook": "<One concise strategic recommendation for the management team. Action-oriented. No post names.>",
    "Instagram": "<One concise strategic recommendation. Action-oriented. No post names.>",
    "TikTok": "<One concise strategic recommendation. Action-oriented. No post names.>",
    "YouTube": "<One concise strategic recommendation. Action-oriented. No post names.>",
    "LinkedIn": "<One concise strategic recommendation. Action-oriented. No post names.>"
  }},
  "period": "{period_label}"
}}

Strict rules you must follow:
1. key_highlights must have EXACTLY 2 items.
2. Point 1 = #1 platform by Avg ER. Point 2 = #2 platform by Avg ER.
3. Bold platform names in highlights using **PlatformName** markdown.
4. NEVER mention any individual post title or post name anywhere in the entire response.
5. NEVER mention the content type "General / Other".
6. Only include in recommendations the platforms that appear in the data.
7. Write for a senior management audience — be concise, strategic, and data-grounded.
8. Use real numbers from the data (ER%, engagement counts, reach)."""

    # ── Call Groq (free, no credit card needed) ─────────────────────────────────
    GROQ_URL = "https://api.groq.com/openai/v1/chat/completions"
    # Try models in order — all free on Groq
    GROQ_MODELS = [
        "llama-3.3-70b-versatile",
        "llama-3.1-70b-versatile",
        "mixtral-8x7b-32768",
        "llama3-70b-8192",
    ]
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    last_error = None
    for model_name in GROQ_MODELS:
        try:
            payload = {
                "model": model_name,
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.3,
                "max_tokens": 2048,
            }
            resp = http_requests.post(GROQ_URL, headers=headers, json=payload, timeout=60)
            if resp.status_code == 200:
                raw = resp.json()["choices"][0]["message"]["content"].strip()
                # Strip markdown fences if present
                if raw.startswith("```"):
                    raw = raw.split("\n", 1)[-1]
                    raw = raw.rsplit("```", 1)[0].strip()
                return json.loads(raw)
            elif resp.status_code in (404, 400):
                last_error = resp.text
                continue  # try next model
            else:
                return {"error": f"Groq API error {resp.status_code}: {resp.text}"}
        except json.JSONDecodeError as e:
            return {"error": f"AI returned invalid JSON: {str(e)}"}
        except Exception as e:
            last_error = str(e)
            continue

    return {"error": f"No Groq model succeeded. Last error: {last_error}"}


# ── Helper: call Groq with a prompt ────────────────────────────────────────────
def _call_groq(api_key: str, prompt: str, max_tokens: int = 3000):
    GROQ_URL = "https://api.groq.com/openai/v1/chat/completions"
    GROQ_MODELS = ["llama-3.3-70b-versatile", "llama-3.1-70b-versatile",
                   "llama3-70b-8192", "mixtral-8x7b-32768"]
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    last_error = None
    for model in GROQ_MODELS:
        try:
            payload = {"model": model, "messages": [{"role": "user", "content": prompt}],
                       "temperature": 0.35, "max_tokens": max_tokens}
            resp = http_requests.post(GROQ_URL, headers=headers, json=payload, timeout=90)
            if resp.status_code == 200:
                raw = resp.json()["choices"][0]["message"]["content"].strip()
                if raw.startswith("```"):
                    raw = raw.split("\n", 1)[-1].rsplit("```", 1)[0].strip()
                return json.loads(raw)
            elif resp.status_code in (404, 400):
                last_error = resp.text
                continue
            else:
                return {"error": f"Groq API error {resp.status_code}: {resp.text}"}
        except json.JSONDecodeError as e:
            return {"error": f"AI returned invalid JSON: {e}"}
        except Exception as e:
            last_error = str(e)
            continue
    return {"error": f"No Groq model succeeded. Last error: {last_error}"}


@app.get("/api/strategy-insights")
def get_strategy_insights(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None)
):
    """Generate Platform Strategy Recommendations and Key Learnings using AI."""
    api_key = os.getenv("GROQ_API_KEY", "")
    if not api_key:
        return {"error": "GROQ_API_KEY not set."}

    df = get_filtered_data(start_date, end_date)
    if df.empty:
        return {"error": "No data for the selected period."}

    PLATS = ["Facebook", "Instagram", "TikTok", "YouTube", "LinkedIn"]
    period_label = f"{start_date or 'the beginning'} to {end_date or 'present'}"

    def _r(v, n=2): return round(float(v), n) if v is not None else 0

    # ── Build format-level breakdown per platform ───────────────────────────────
    plat_stats = {}
    for plat in PLATS:
        pdf = df[df["platform"] == plat]
        if pdf.empty:
            continue
        org  = pdf[pdf["is_organic"] == True]
        paid = pdf[pdf["is_organic"] == False]

        org_er  = _r(org["engagement_rate"].mean())  if not org.empty  else 0
        paid_er = _r(paid["engagement_rate"].mean()) if not paid.empty else 0
        all_er  = _r(pdf["engagement_rate"].mean())

        # Format breakdown — ALL content (organic + paid): count, avg ER, best ER, worst ER
        fmt_rows = []
        if "format" in pdf.columns:
            for fmt, grp in pdf.groupby("format"):
                fmt = str(fmt).strip()
                if not fmt or fmt == "nan" or len(grp) == 0:
                    continue
                fmt_rows.append({
                    "format": fmt,
                    "posts": len(grp),
                    "avg_er": _r(grp["engagement_rate"].mean()),
                    "best_er": _r(grp["engagement_rate"].max()),
                    "worst_er": _r(grp["engagement_rate"].min()),
                })
        fmt_rows.sort(key=lambda x: x["avg_er"], reverse=True)

        er_spread = _r(pdf["engagement_rate"].max() - pdf["engagement_rate"].min()) if not pdf.empty else 0

        # Top 5 and bottom 5 posts by ER (ALL content) — truncated title for theme inference
        title_col = "title" if "title" in pdf.columns else None
        def _title(row):
            t = str(row.get(title_col, "") or "").strip() if title_col else ""
            return t[:80] if t and t != "nan" else "(no title)"

        top5 = pdf.sort_values("engagement_rate", ascending=False).head(5)
        bot5 = pdf[pdf["engagement_rate"] > 0].sort_values("engagement_rate").head(5)
        top5_posts = [{"er": _r(r["engagement_rate"]), "title": _title(r)} for _, r in top5.iterrows()]
        bot5_posts = [{"er": _r(r["engagement_rate"]), "title": _title(r)} for _, r in bot5.iterrows()]

        plat_stats[plat] = {
            "total_posts":        len(pdf),
            "organic_posts":      len(org),
            "paid_posts":         len(paid),
            "avg_er_all":         all_er,
            "avg_er_organic":     org_er,
            "avg_er_paid":        paid_er,
            "avg_reach":          int(pdf["reach"].mean()) if not pdf.empty else 0,
            "er_spread":          er_spread,
            "format_performance": fmt_rows,        # ALL content
            "top5_posts":         top5_posts,      # ALL content, for theme inference
            "bot5_posts":         bot5_posts,      # ALL content, for theme inference
        }

    er_rank = sorted(plat_stats.items(), key=lambda x: x[1]["avg_er_all"], reverse=True)
    er_rank_str = ", ".join([f"{p} ({d['avg_er_all']}% avg ER, {d['total_posts']} total posts, {d['organic_posts']} organic)" for p, d in er_rank])

    data_str = json.dumps(plat_stats, indent=2)


    # ── Prompt 1: Platform Strategy ─────────────────────────────────────────────
    prompt1 = f"""You are a senior social media strategist producing a board-level report for CIMB Bank Malaysia.
Period: {period_label}
Platform organic ER ranking: {er_rank_str}

Per-platform analytics (format breakdown + top/bottom organic post titles for theme context):
{data_str}

TASK: Write a STOP / PAUSE / CONTINUE / ENHANCE strategy for each platform.

HOW TO READ THE DATA:
- "top5_organic_posts" gives you the titles of the BEST performing content — use these to infer CONTENT THEMES (e.g. "Security/fraud awareness", "Financial literacy tips", "Lifestyle/festive content", "Product promotions", "Community stories", "Employer brand")
- "bot5_organic_posts" gives you the WORST performing content — use these to infer what themes/approaches to STOP or PAUSE
- "format_performance" tells you which FORMATS (Reel, Video, Static, Carousel, etc.) drove the highest vs lowest ER

HOW TO WRITE EACH CELL:
- STOP: The content THEMES or APPROACHES that consistently underperformed. Infer from bot5 titles + lowest ER formats. Be specific — e.g. "Static product announcement posts" or "Generic promotional announcements"
- PAUSE: Themes or formats that showed inconsistent results (high er_spread) — currently working sometimes but need refinement. E.g. "Carousel-format campaign posts" or "Paid promotional reels"
- CONTINUE: The THEMES and FORMATS with proven consistently high ER. Infer from top5 titles. E.g. "Security/fraud awareness content", "Financial literacy explainers", "Creator-led storytelling reels"
- ENHANCE: High-potential themes/formats that appeared in top performers but are underutilised (few posts, high best_er). E.g. "Collaborations with creators", "Series-based educational content", "UGC-style community stories"

Each cell = SHORT PHRASES only (max 15 words). Can list 2-3 themes separated by commas if relevant. NOT full sentences.

Return ONLY valid JSON, no markdown, no code fences:
{{
  "platform_strategy": {{
    "Facebook":  {{"stop": "...", "pause": "...", "continue": "...", "enhance": "..."}},
    "Instagram": {{"stop": "...", "pause": "...", "continue": "...", "enhance": "..."}},
    "TikTok":    {{"stop": "...", "pause": "...", "continue": "...", "enhance": "..."}},
    "YouTube":   {{"stop": "...", "pause": "...", "continue": "...", "enhance": "..."}},
    "LinkedIn":  {{"stop": "...", "pause": "...", "continue": "...", "enhance": "..."}}
  }},
  "key_takeaways": ["...", "...", "...", "...", "..."]
}}

Additional rules:
- If a platform has no clear underperformer, write "No change, to monitor performance" for STOP.
- If no clear PAUSE needed, write "No pause required".
- Key takeaways: exactly 5. Each = one professional sentence referencing real ER% numbers AND specific content themes from the data.
- Think like a strategist reading real content — what TOPICS and APPROACHES worked, not just what format file type was used."""

    # ── Prompt 2: Key Learnings ──────────────────────────────────────────────────
    prompt2 = f"""You are a senior social media strategist producing a board-level report for CIMB Bank Malaysia.
Period: {period_label}
Platform organic ER ranking: {er_rank_str}

Per-platform analytics (format breakdown + top/bottom organic post titles):
{data_str}

TASK: Write 4 KEY LEARNINGS that describe CROSS-PLATFORM patterns.

CRITICAL RULE: Each learning MUST observe a pattern that appears ACROSS MULTIPLE PLATFORMS — not a single platform's performance. The insight should be something the ENTIRE content team can act on regardless of which platform they manage.

Good examples of cross-platform learnings (from real reports):
- "Episodic finance advice is a repeatable engagement driver" → TikTok AND Instagram AND YouTube all showed high ER for financial education content delivered in a series format
- "Interactive or participatory content mechanics boost comments and shares" → Multiple platforms benefited from posts that asked audiences to do something (pause, comment, choose)
- "Promotional content needs utility or a mechanic to work" → Generic product posts underperformed across Facebook, Instagram and TikTok unless they had a reward, challenge or explainer element
- "Cultural and festive content creates spikes when personalised" → Raya/festive content spiked ER on both TikTok and Instagram when it was specific and relatable

HOW TO USE THE DATA:
- Look at top5_organic_posts across ALL platforms — what COMMON THEMES appear in the highest-performing content?
- Look at bot5_organic_posts across ALL platforms — what COMMON WEAKNESSES appear?
- Look at format_performance across platforms — are there formats that consistently outperform or underperform?
- Use real ER% numbers from the data to support each insight

STRUCTURE per learning:
- title: Short punchy phrase (the insight in 8 words or less)
- description: 2-3 sentences. Reference 2+ platforms. Include real ER% numbers. Explain WHY the pattern works.
- action: One concrete directive for the content team. Start with a strong verb. Apply across platforms.

Return ONLY valid JSON, no markdown, no code fences:
{{
  "key_learnings": [
    {{"number": 1, "title": "...", "description": "...", "action": "..."}},
    {{"number": 2, "title": "...", "description": "...", "action": "..."}},
    {{"number": 3, "title": "...", "description": "...", "action": "..."}},
    {{"number": 4, "title": "...", "description": "...", "action": "..."}}
  ]
}}

Additional rules:
- Exactly 4 learnings, each on a DIFFERENT cross-platform theme.
- Vary the topics: content theme patterns, format effectiveness, audience engagement mechanics, organic vs paid behaviour.
- DO NOT write a learning that only talks about one platform.
- DO NOT mention specific post titles — infer the content theme and describe it generically."""

    # ── Call Groq (with pause between calls to avoid free-tier TPM rate limit) ───
    import time
    strategy_result = _call_groq(api_key, prompt1, max_tokens=2500)
    if "error" in strategy_result and "429" in str(strategy_result.get("error", "")):
        return {"error": "Groq rate limit hit on first call. Please wait 30 seconds and try again."}
    time.sleep(15)   # wait 15s so TPM window resets before second call
    learnings_result = _call_groq(api_key, prompt2, max_tokens=2500)
    if "error" in learnings_result and "429" in str(learnings_result.get("error", "")):
        return {"error": "Groq rate limit hit on second call. Please wait 30 seconds and try again."}

    if "error" in strategy_result:
        return {"error": f"Strategy prompt failed: {strategy_result['error']}"}
    if "error" in learnings_result:
        return {"error": f"Learnings prompt failed: {learnings_result['error']}"}

    return {
        "period":            period_label,
        "platform_strategy": strategy_result.get("platform_strategy", {}),
        "key_takeaways":     strategy_result.get("key_takeaways", []),
        "key_learnings":     learnings_result.get("key_learnings", []),
    }



