#!/usr/bin/env python3
"""Build a Metricool Instagram/TikTok per-post engagement-rate proof workbook.

Formula:
    Engagement Rate = (likes + comments + shares + saves) / reach

TikTok note:
    Metricool's TikTok post endpoint exposes viewCount rather than reach, so
    this script uses viewCount as the reach proxy for TikTok.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_URL = "https://app.metricool.com/api"
DEFAULT_USER_ID = "4327762"


Row = Dict[str, Any]


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def load_metricool_env() -> None:
    script_path = Path(__file__).resolve()
    candidates = [
        Path.cwd() / ".env",
        Path.cwd() / "metricool" / ".env",
        script_path.parents[1] / ".env",  # skill dir
        script_path.parents[2] / ".env",  # repo root, usually
        script_path.parents[2] / "metricool" / ".env",
    ]
    for path in candidates:
        load_env_file(path)


def normalize_date(value: str, end_of_day: bool = False) -> str:
    value = value.strip()
    if value.lower() == "today":
        date = datetime.now().strftime("%Y-%m-%d")
        return f"{date}T{'23:59:59' if end_of_day else '00:00:00'}"
    if "T" in value:
        return value
    # Validate basic YYYY-MM-DD input.
    datetime.strptime(value, "%Y-%m-%d")
    return f"{value}T{'23:59:59' if end_of_day else '00:00:00'}"


def safe_filename(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return re.sub(r"_+", "_", value).strip("_") or "metricool_report"


def safe_num(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def clean_text(value: Any) -> Any:
    # Excel XML cannot store lone surrogate characters. Replace invalid codepoints.
    if isinstance(value, str):
        return value.encode("utf-8", "replace").decode("utf-8", "replace")
    return value


def clean_row(values: Iterable[Any]) -> List[Any]:
    return [clean_text(v) for v in values]


class MetricoolClient:
    def __init__(self, blog_id: str, token: str, user_id: str = DEFAULT_USER_ID):
        self.blog_id = blog_id
        self.token = token
        self.user_id = user_id

    def get(self, endpoint: str, params: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        params = params or {}
        query = {
            "userId": self.user_id,
            "userToken": self.token,
            "blogId": self.blog_id,
            **params,
        }
        url = f"{BASE_URL}{endpoint}?{urllib.parse.urlencode(query)}"
        request = urllib.request.Request(url, headers={"Accept": "application/json"})
        try:
            with urllib.request.urlopen(request, timeout=60) as response:
                body = response.read().decode("utf-8")
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")[:500]
            raise RuntimeError(f"Metricool API error {exc.code} for {endpoint}: {detail}") from exc
        if not body.strip():
            return {}
        return json.loads(body)

    def brand_name(self) -> Optional[str]:
        try:
            data = self.get(f"/v2/settings/brands/{self.blog_id}").get("data", {})
            return data.get("label") or data.get("name")
        except Exception:
            return None

    def fetch_paginated(
        self,
        endpoint: str,
        from_iso: str,
        to_iso: str,
        raw_dir: Optional[Path],
        raw_name: str,
        size: int = 100,
    ) -> List[Dict[str, Any]]:
        items: List[Dict[str, Any]] = []
        seen = set()
        page = 0
        while True:
            obj = self.get(endpoint, {"from": from_iso, "to": to_iso, "page": page, "size": size})
            if raw_dir:
                raw_dir.mkdir(parents=True, exist_ok=True)
                (raw_dir / f"{raw_name}_page_{page}.json").write_text(
                    json.dumps(obj, indent=2, ensure_ascii=False), encoding="utf-8"
                )
            data = obj.get("data", []) if isinstance(obj, dict) else obj
            if not data:
                break
            new_count = 0
            for item in data:
                key = (
                    item.get("postId")
                    or item.get("reelId")
                    or item.get("videoId")
                    or json.dumps(item, sort_keys=True, default=str)
                )
                if key not in seen:
                    seen.add(key)
                    items.append(item)
                    new_count += 1
            # Stop if API ignores pagination and repeats results, or if final short page.
            if new_count == 0 or len(data) < size:
                break
            page += 1
            time.sleep(0.2)
        if raw_dir:
            (raw_dir / f"{raw_name}_combined.json").write_text(
                json.dumps({"data": items}, indent=2, ensure_ascii=False), encoding="utf-8"
            )
        return items


def instagram_rows(items: Iterable[Dict[str, Any]], post_format: str) -> List[Row]:
    rows: List[Row] = []
    for post in items:
        likes = safe_num(post.get("likes"))
        comments = safe_num(post.get("comments"))
        shares = safe_num(post.get("shares"))
        saved = safe_num(post.get("saved"))
        engagements = likes + comments + shares + saved
        reach = safe_num(post.get("reach"))
        er = engagements / reach if reach else None
        metricool_er = post.get("engagement")
        published = post.get("publishedAt") or {}
        rows.append(
            {
                "network": "Instagram",
                "format": post_format,
                "post_id": post.get("postId") or post.get("reelId"),
                "published_at": published.get("dateTime"),
                "timezone": published.get("timezone"),
                "type": post.get("type"),
                "url": post.get("url"),
                "caption": post.get("content"),
                "likes": likes,
                "comments": comments,
                "shares": shares,
                "saved": saved,
                "engagements": engagements,
                "denominator_label": "Reach",
                "reach_or_views": reach,
                "engagement_rate": er,
                "metricool_engagement_rate": safe_num(metricool_er) / 100 if metricool_er is not None else None,
                "impressions": safe_num(post.get("impressionsTotal")),
                "views": safe_num(post.get("views")),
            }
        )
    return rows


def tiktok_rows(items: Iterable[Dict[str, Any]]) -> List[Row]:
    rows: List[Row] = []
    for post in items:
        likes = safe_num(post.get("likeCount"))
        comments = safe_num(post.get("commentCount"))
        shares = safe_num(post.get("shareCount"))
        saved = 0.0
        engagements = likes + comments + shares + saved
        views = safe_num(post.get("viewCount"))
        er = engagements / views if views else None
        metricool_er = post.get("engagement")
        rows.append(
            {
                "network": "TikTok",
                "format": "Video",
                "post_id": post.get("videoId"),
                "published_at": post.get("createTime"),
                "timezone": "",
                "type": post.get("type"),
                "url": post.get("shareUrl"),
                "caption": post.get("videoDescription") or post.get("title"),
                "likes": likes,
                "comments": comments,
                "shares": shares,
                "saved": saved,
                "engagements": engagements,
                "denominator_label": "Views (TikTok reach proxy from Metricool)",
                "reach_or_views": views,
                "engagement_rate": er,
                "metricool_engagement_rate": safe_num(metricool_er) / 100 if metricool_er is not None else None,
                "impressions": "",
                "views": views,
            }
        )
    return rows


def summarize(rows: List[Row]) -> Dict[str, Any]:
    valid = [row for row in rows if row["engagement_rate"] is not None]
    total_engagements = sum(safe_num(row["engagements"]) for row in rows)
    total_denominator = sum(safe_num(row["reach_or_views"]) for row in rows)
    return {
        "posts": len(rows),
        "valid_posts": len(valid),
        "total_engagements": total_engagements,
        "total_denominator": total_denominator,
        "avg_er": mean([row["engagement_rate"] for row in valid]) if valid else None,
        "weighted_er": total_engagements / total_denominator if total_denominator else None,
    }


def build_workbook(
    output: Path,
    brand: str,
    blog_id: str,
    from_iso: str,
    to_iso: str,
    instagram_all: List[Row],
    tiktok: List[Row],
    ig_feed: List[Row],
    ig_reels: List[Row],
) -> Dict[str, Dict[str, Any]]:
    headers = [
        "Network",
        "Format",
        "Post ID",
        "Published At",
        "Timezone",
        "Type",
        "URL",
        "Caption",
        "Likes",
        "Comments",
        "Shares",
        "Saved",
        "Total Engagements",
        "Denominator",
        "Reach / Views",
        "Engagement Rate",
        "Metricool ER",
        "Impressions",
        "Views",
        "Proof Formula",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(row):
        for cell in row:
            cell.fill = header_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    def add_proof_sheet(title: str, rows: List[Row]) -> None:
        sheet = wb.create_sheet(title)
        sheet.append(headers)
        style_header(sheet[1])
        for row in rows:
            sheet.append(
                clean_row(
                    [
                        row["network"],
                        row["format"],
                        row["post_id"],
                        row["published_at"],
                        row["timezone"],
                        row["type"],
                        row["url"],
                        row["caption"],
                        row["likes"],
                        row["comments"],
                        row["shares"],
                        row["saved"],
                        row["engagements"],
                        row["denominator_label"],
                        row["reach_or_views"],
                        row["engagement_rate"],
                        row["metricool_engagement_rate"],
                        row["impressions"],
                        row["views"],
                        "=(Likes+Comments+Shares+Saved)/(Reach or Views)",
                    ]
                )
            )
        for xl_row in sheet.iter_rows(min_row=2):
            for cell in xl_row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(cell.column in [7, 8, 20]))
        for col in [16, 17]:
            for cells in sheet.iter_cols(min_col=col, max_col=col, min_row=2, max_row=sheet.max_row):
                for cell in cells:
                    if cell.value is not None:
                        cell.number_format = "0.00%"
        for col in list(range(9, 16)) + [18, 19]:
            for cells in sheet.iter_cols(min_col=col, max_col=col, min_row=2, max_row=sheet.max_row):
                for cell in cells:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0"
        widths = {
            1: 14,
            2: 14,
            3: 28,
            4: 24,
            5: 16,
            6: 22,
            7: 42,
            8: 65,
            9: 10,
            10: 10,
            11: 10,
            12: 10,
            13: 18,
            14: 36,
            15: 15,
            16: 17,
            17: 14,
            18: 14,
            19: 12,
            20: 42,
        }
        for idx, width in widths.items():
            sheet.column_dimensions[get_column_letter(idx)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    add_proof_sheet("Instagram Proof", instagram_all)
    add_proof_sheet("TikTok Proof", tiktok)
    add_proof_sheet("IG Feed Raw Proof", ig_feed)
    add_proof_sheet("IG Reels Raw Proof", ig_reels)

    summaries = {
        "Instagram Feed": summarize(ig_feed),
        "Instagram Reels": summarize(ig_reels),
        "Instagram Total": summarize(instagram_all),
        "TikTok": summarize(tiktok),
    }

    ws["A1"] = f"{brand} — Instagram & TikTok Engagement Rate Proof"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "Brand"
    ws["B3"] = brand
    ws["A4"] = "Blog ID"
    ws["B4"] = blog_id
    ws["A5"] = "Source"
    ws["B5"] = "Metricool API v2 analytics post endpoints"
    ws["A6"] = "Date range"
    ws["B6"] = f"{from_iso} to {to_iso}"
    ws["A7"] = "Calculation"
    ws["B7"] = (
        "Engagement Rate = (Likes + Comments + Shares + Saved) / Reach. "
        "For TikTok, Metricool exposes viewCount rather than reach, so viewCount is used "
        "as the denominator/reach proxy. Rows with denominator 0 are excluded from average ER."
    )
    ws["B7"].alignment = Alignment(wrap_text=True, vertical="top")

    for cell in ws["A3:A7"]:
        cell[0].font = bold_font

    start = 10
    summary_headers = [
        "Network / Format",
        "Posts",
        "Posts with denominator",
        "Total Engagements",
        "Total Reach / Views",
        "Average ER",
        "Weighted ER",
    ]
    ws.append([])
    ws.append(summary_headers)
    style_header(ws[start])
    for label, stats in summaries.items():
        ws.append(
            [
                label,
                stats["posts"],
                stats["valid_posts"],
                stats["total_engagements"],
                stats["total_denominator"],
                stats["avg_er"],
                stats["weighted_er"],
            ]
        )

    for row in ws.iter_rows(min_row=start + 1, max_row=start + len(summaries)):
        for cell in row:
            cell.border = border
        row[5].number_format = "0.00%"
        row[6].number_format = "0.00%"
        row[3].number_format = "#,##0"
        row[4].number_format = "#,##0"

    for col, width in {1: 24, 2: 34, 3: 24, 4: 20, 5: 20, 6: 14, 7: 14}.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[7].height = 48
    ws.freeze_panes = "A10"
    ws["A17"] = "Tabs"
    ws["A17"].font = bold_font
    ws["B17"] = (
        "Instagram Proof combines feed + reels. Separate IG Feed Raw Proof and IG Reels Raw Proof "
        "tabs are included for auditability. TikTok Proof contains TikTok videos."
    )
    ws["B17"].alignment = Alignment(wrap_text=True)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return summaries


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build Metricool engagement-rate proof XLSX for Instagram/TikTok.")
    parser.add_argument("--blog-id", required=True, help="Metricool blogId/brand ID")
    parser.add_argument("--brand", help="Brand name for the workbook. If omitted, fetched from Metricool when possible.")
    parser.add_argument("--from", dest="from_date", required=True, help="Start date: YYYY-MM-DD, ISO datetime, or today")
    parser.add_argument("--to", dest="to_date", required=True, help="End date: YYYY-MM-DD, ISO datetime, or today")
    parser.add_argument("--networks", default="instagram,tiktok", help="Comma-separated networks: instagram,tiktok")
    parser.add_argument("--output", help="Output XLSX path")
    parser.add_argument("--raw-dir", help="Directory for raw JSON responses")
    parser.add_argument("--no-raw", action="store_true", help="Do not save raw Metricool JSON responses")
    parser.add_argument("--no-instagram-reels", action="store_true", help="Do not fetch Instagram reels")
    parser.add_argument("--user-id", default=os.environ.get("METRICOOL_USER_ID", DEFAULT_USER_ID), help="Metricool userId")
    return parser.parse_args()


def main() -> None:
    load_metricool_env()
    args = parse_args()
    token = os.environ.get("METRICOOL_TOKEN", "")
    if not token:
        raise SystemExit("METRICOOL_TOKEN is not set. Run metricool setup or export METRICOOL_TOKEN.")

    from_iso = normalize_date(args.from_date, end_of_day=False)
    to_iso = normalize_date(args.to_date, end_of_day=True)
    from_label = from_iso.split("T", 1)[0]
    to_label = to_iso.split("T", 1)[0]

    client = MetricoolClient(args.blog_id, token, args.user_id)
    brand = args.brand or client.brand_name() or f"Metricool Brand {args.blog_id}"

    if args.output:
        output = Path(args.output)
    else:
        output = Path(
            f"{safe_filename(brand)}_Instagram_TikTok_Engagement_Rate_{from_label}_to_{to_label}.xlsx"
        )

    raw_dir: Optional[Path]
    if args.no_raw:
        raw_dir = None
    elif args.raw_dir:
        raw_dir = Path(args.raw_dir)
    else:
        raw_dir = Path(f"{safe_filename(brand)}_metricool_raw_{from_label}_to_{to_label}")

    networks = {part.strip().lower() for part in args.networks.split(",") if part.strip()}

    ig_feed: List[Row] = []
    ig_reels: List[Row] = []
    tiktok: List[Row] = []

    if "instagram" in networks:
        print("Fetching Instagram feed posts...")
        feed_raw = client.fetch_paginated(
            "/v2/analytics/posts/instagram", from_iso, to_iso, raw_dir, "instagram_feed"
        )
        ig_feed = instagram_rows(feed_raw, "Feed Post")
        if not args.no_instagram_reels:
            print("Fetching Instagram reels...")
            reels_raw = client.fetch_paginated(
                "/v2/analytics/reels/instagram", from_iso, to_iso, raw_dir, "instagram_reels"
            )
            ig_reels = instagram_rows(reels_raw, "Reel")

    if "tiktok" in networks:
        print("Fetching TikTok posts...")
        tiktok_raw = client.fetch_paginated("/v2/analytics/posts/tiktok", from_iso, to_iso, raw_dir, "tiktok")
        tiktok = tiktok_rows(tiktok_raw)

    instagram_all = sorted(ig_feed + ig_reels, key=lambda row: row.get("published_at") or "", reverse=True)
    tiktok = sorted(tiktok, key=lambda row: row.get("published_at") or "", reverse=True)

    summaries = build_workbook(output, brand, args.blog_id, from_iso, to_iso, instagram_all, tiktok, ig_feed, ig_reels)

    print(f"Saved: {output}")
    if raw_dir:
        print(f"Raw JSON: {raw_dir}")
    for label, stats in summaries.items():
        avg = stats["avg_er"]
        weighted = stats["weighted_er"]
        avg_s = f"{avg:.2%}" if avg is not None else "n/a"
        weighted_s = f"{weighted:.2%}" if weighted is not None else "n/a"
        print(
            f"{label}: posts={stats['posts']}, valid={stats['valid_posts']}, "
            f"avg_er={avg_s}, weighted_er={weighted_s}"
        )


if __name__ == "__main__":
    main()
