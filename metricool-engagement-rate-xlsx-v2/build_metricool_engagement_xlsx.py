#!/usr/bin/env python3
"""Build a Metricool Instagram / TikTok / YouTube / LinkedIn per-post
engagement-rate proof workbook.

Formula:
    Engagement Rate = (likes + comments + shares + saves) / reach

Platform notes:
    Instagram : uses Metricool `reach` as the denominator.
    TikTok    : Metricool exposes `viewCount`, not reach — used as proxy.
                NOTE: Metricool's TikTok API does not return a Saves/Favorites
                field. Saves are therefore excluded from TikTok ER; a warning
                is written to the Summary sheet automatically.
    YouTube   : uses `viewCount` as the denominator (reach not exposed).
                Engagements = likes + comments + shares.
    LinkedIn  : uses `impressions` as the denominator when reach is unavailable.
                Engagements = likes + comments + shares + clicks.

Min-views filter (--min-views N):
    TikTok posts with N views/reach or fewer are excluded from average ER
    calculation. They still appear in proof tabs but are marked as excluded.
    Default: 0 (no filter). Recommended: 20.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_URL = "https://app.metricool.com/api"
DEFAULT_USER_ID = "4327762"

# TikTok Saves limitation notice shown in Summary sheet
TIKTOK_SAVES_NOTE = (
    "⚠️  SAVES DATA UNAVAILABLE FOR TIKTOK: Metricool's TikTok API endpoint "
    "does not return a Saves / Add-to-Favorites field. TikTok ER in this "
    "workbook is calculated as (Likes + Comments + Shares) / Views only. "
    "Actual ER from TikTok native analytics will be slightly higher because "
    "it includes Saves. Use TikTok native export for the most accurate ER."
)

Row = Dict[str, Any]


# ── Environment helpers ────────────────────────────────────────────────────────

def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


def load_metricool_env() -> None:
    script_path = Path(__file__).resolve()
    candidates = [
        Path.cwd() / ".env",
        Path.cwd() / "metricool" / ".env",
        script_path.parents[1] / ".env",
        script_path.parents[2] / ".env",
        script_path.parents[2] / "metricool" / ".env",
    ]
    for path in candidates:
        load_env_file(path)


# ── Utility helpers ────────────────────────────────────────────────────────────

def normalize_date(value: str, end_of_day: bool = False) -> str:
    value = value.strip()
    if value.lower() == "today":
        date = datetime.now().strftime("%Y-%m-%d")
        return f"{date}T{'23:59:59' if end_of_day else '00:00:00'}"
    if "T" in value:
        return value
    datetime.strptime(value, "%Y-%m-%d")
    return f"{value}T{'23:59:59' if end_of_day else '00:00:00'}"


def safe_filename(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return re.sub(r"_+", "_", value).strip("_") or "metricool_report"


def safe_num(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def clean_text(value: Any) -> Any:
    if isinstance(value, str):
        return value.encode("utf-8", "replace").decode("utf-8", "replace")
    return value


def clean_row(values: Iterable[Any]) -> List[Any]:
    return [clean_text(v) for v in values]


# ── Metricool API client ───────────────────────────────────────────────────────

class MetricoolClient:
    def __init__(self, blog_id: str, token: str, user_id: str = DEFAULT_USER_ID):
        self.blog_id = blog_id
        self.token = token
        self.user_id = user_id

    def get(self, endpoint: str, params: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        params = params or {}
        query = {
            "userId": self.user_id,
            "userToken": self.token,
            "blogId": self.blog_id,
            **params,
        }
        url = f"{BASE_URL}{endpoint}?{urllib.parse.urlencode(query)}"
        request = urllib.request.Request(url, headers={"Accept": "application/json"})
        try:
            with urllib.request.urlopen(request, timeout=60) as response:
                body = response.read().decode("utf-8")
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")[:500]
            raise RuntimeError(f"Metricool API error {exc.code} for {endpoint}: {detail}") from exc
        if not body.strip():
            return {}
        return json.loads(body)

    def brand_name(self) -> Optional[str]:
        try:
            data = self.get(f"/v2/settings/brands/{self.blog_id}").get("data", {})
            return data.get("label") or data.get("name")
        except Exception:
            return None

    def fetch_paginated(
        self,
        endpoint: str,
        from_iso: str,
        to_iso: str,
        raw_dir: Optional[Path],
        raw_name: str,
        size: int = 100,
    ) -> List[Dict[str, Any]]:
        items: List[Dict[str, Any]] = []
        seen: set = set()
        page = 0
        while True:
            obj = self.get(endpoint, {"from": from_iso, "to": to_iso, "page": page, "size": size})
            if raw_dir:
                raw_dir.mkdir(parents=True, exist_ok=True)
                (raw_dir / f"{raw_name}_page_{page}.json").write_text(
                    json.dumps(obj, indent=2, ensure_ascii=False), encoding="utf-8"
                )
            data = obj.get("data", []) if isinstance(obj, dict) else obj
            if not data:
                break
            new_count = 0
            for item in data:
                key = (
                    item.get("postId")
                    or item.get("reelId")
                    or item.get("videoId")
                    or item.get("youtubeVideoId")
                    or item.get("linkedinPostId")
                    or json.dumps(item, sort_keys=True, default=str)
                )
                if key not in seen:
                    seen.add(key)
                    items.append(item)
                    new_count += 1
            if new_count == 0 or len(data) < size:
                break
            page += 1
            time.sleep(0.2)
        if raw_dir:
            (raw_dir / f"{raw_name}_combined.json").write_text(
                json.dumps({"data": items}, indent=2, ensure_ascii=False), encoding="utf-8"
            )
        return items


# ── Per-platform row builders ──────────────────────────────────────────────────

def instagram_rows(items: Iterable[Dict[str, Any]], post_format: str) -> List[Row]:
    rows: List[Row] = []
    for post in items:
        likes = safe_num(post.get("likes"))
        comments = safe_num(post.get("comments"))
        shares = safe_num(post.get("shares"))
        saved = safe_num(post.get("saved"))
        engagements = likes + comments + shares + saved
        reach = safe_num(post.get("reach"))
        er = engagements / reach if reach else None
        metricool_er = post.get("engagement")
        published = post.get("publishedAt") or {}
        rows.append(
            {
                "network": "Instagram",
                "format": post_format,
                "post_id": post.get("postId") or post.get("reelId"),
                "published_at": published.get("dateTime"),
                "timezone": published.get("timezone"),
                "type": post.get("type"),
                "url": post.get("url"),
                "caption": post.get("content"),
                "likes": likes,
                "comments": comments,
                "shares": shares,
                "saved": saved,
                "engagements": engagements,
                "denominator_label": "Reach",
                "reach_or_views": reach,
                "engagement_rate": er,
                "metricool_engagement_rate": safe_num(metricool_er) / 100 if metricool_er is not None else None,
                "impressions": safe_num(post.get("impressionsTotal")),
                "views": safe_num(post.get("views")),
                "saves_available": True,
            }
        )
    return rows


def tiktok_rows(items: Iterable[Dict[str, Any]]) -> List[Row]:
    """Build TikTok rows.

    NOTE: Metricool's TikTok API does not expose a Saves/Favorites field.
    The raw post object contains only: likeCount, commentCount, shareCount,
    viewCount, engagement, impressionSources. Saves are set to 0.0 and a
    workbook-level warning is added in the Summary sheet.
    """
    rows: List[Row] = []
    for post in items:
        likes = safe_num(post.get("likeCount"))
        comments = safe_num(post.get("commentCount"))
        shares = safe_num(post.get("shareCount"))
        # Saves not available from Metricool TikTok API — confirmed missing field.
        saved = 0.0
        engagements = likes + comments + shares  # saves excluded intentionally
        views = safe_num(post.get("viewCount"))
        er = engagements / views if views else None
        metricool_er = post.get("engagement")
        rows.append(
            {
                "network": "TikTok",
                "format": "Video",
                "post_id": post.get("videoId"),
                "published_at": post.get("createTime"),
                "timezone": "",
                "type": post.get("type"),
                "url": post.get("shareUrl"),
                "caption": post.get("videoDescription") or post.get("title"),
                "likes": likes,
                "comments": comments,
                "shares": shares,
                "saved": saved,
                "engagements": engagements,
                "denominator_label": "Views (TikTok reach proxy — Saves excluded, see Summary note)",
                "reach_or_views": views,
                "engagement_rate": er,
                "metricool_engagement_rate": safe_num(metricool_er) / 100 if metricool_er is not None else None,
                "impressions": "",
                "views": views,
                "saves_available": False,
            }
        )
    return rows


def youtube_rows(items: Iterable[Dict[str, Any]]) -> List[Row]:
    """Build YouTube rows.

    Metricool YouTube endpoint fields (typical):
        videoId / youtubeVideoId, title, description, publishedAt,
        viewCount, likeCount, commentCount, shareCount, url, videoType
    ER = (likes + comments + shares) / views  (reach not exposed for YouTube)

    Note: YouTube Shorts (videoType = 'SHORT') are excluded because the
    native YouTube Studio export does not include Shorts in its regular
    video analytics, so excluding them keeps both datasets comparable.
    """
    rows: List[Row] = []
    shorts_skipped = 0
    for post in items:
        # Skip YouTube Shorts
        if str(post.get("videoType", "")).upper() == "SHORT":
            shorts_skipped += 1
            continue
        likes = safe_num(post.get("likeCount") or post.get("likes"))
        comments = safe_num(post.get("commentCount") or post.get("comments"))
        shares = safe_num(post.get("shareCount") or post.get("shares"))
        saved = 0.0
        engagements = likes + comments + shares
        views = safe_num(post.get("viewCount") or post.get("views"))
        er = engagements / views if views else None
        metricool_er = post.get("engagement")
        published = post.get("publishedAt") or {}
        published_at = published.get("dateTime") if isinstance(published, dict) else published
        rows.append(
            {
                "network": "YouTube",
                "format": "Video",
                "post_id": post.get("youtubeVideoId") or post.get("videoId"),
                "published_at": published_at,
                "timezone": published.get("timezone") if isinstance(published, dict) else "",
                "type": post.get("type", "VIDEO"),
                "url": post.get("watchUrl") or post.get("url") or post.get("shareUrl"),
                "caption": post.get("title") or post.get("description"),
                "likes": likes,
                "comments": comments,
                "shares": shares,
                "saved": saved,
                "engagements": engagements,
                "denominator_label": "Views (YouTube reach proxy)",
                "reach_or_views": views,
                "engagement_rate": er,
                "metricool_engagement_rate": safe_num(metricool_er) / 100 if metricool_er is not None else None,
                "impressions": safe_num(post.get("impressions") or post.get("impressionsTotal")),
                "views": views,
                "saves_available": False,
            }
        )
    if shorts_skipped:
        print(f"  YouTube: {shorts_skipped} Shorts excluded (videoType=SHORT).")
    return rows


def linkedin_rows(items: Iterable[Dict[str, Any]]) -> List[Row]:
    """Build LinkedIn rows.

    Metricool LinkedIn endpoint fields (typical):
        postId / linkedinPostId, content/text, publishedAt,
        likes/likeCount, comments/commentCount, shares/shareCount,
        clicks, impressions, reach, url
    ER = (likes + comments + shares + clicks) / impressions
         (reach used if impressions unavailable)
    """
    rows: List[Row] = []
    for post in items:
        likes = safe_num(post.get("likeCount") or post.get("likes"))
        comments = safe_num(post.get("commentCount") or post.get("comments"))
        shares = safe_num(post.get("shareCount") or post.get("shares"))
        clicks = safe_num(post.get("clicks") or post.get("clickCount"))
        saved = 0.0
        engagements = likes + comments + shares + clicks
        impressions = safe_num(post.get("impressions") or post.get("impressionsTotal"))
        reach = safe_num(post.get("reach"))
        denominator = impressions if impressions else reach
        denominator_label = "Impressions" if impressions else "Reach"
        er = engagements / denominator if denominator else None
        metricool_er = post.get("engagement")
        published = post.get("created") or post.get("publishedAt") or {}
        published_at = published.get("dateTime") if isinstance(published, dict) else published
        rows.append(
            {
                "network": "LinkedIn",
                "format": post.get("type", "Post"),
                "post_id": post.get("linkedinPostId") or post.get("postId"),
                "published_at": published_at,
                "timezone": published.get("timezone") if isinstance(published, dict) else "",
                "type": post.get("type", "POST"),
                "url": post.get("url"),
                "caption": post.get("comment") or post.get("title") or post.get("content") or post.get("text"),
                "likes": likes,
                "comments": comments,
                "shares": shares,
                "saved": clicks,  # repurpose saved column for LinkedIn clicks
                "engagements": engagements,
                "denominator_label": f"{denominator_label} (LinkedIn — Clicks included in engagements)",
                "reach_or_views": denominator,
                "engagement_rate": er,
                "metricool_engagement_rate": safe_num(metricool_er) / 100 if metricool_er is not None else None,
                "impressions": impressions,
                "views": safe_num(post.get("views") or post.get("videoViews")),
                "saves_available": False,
            }
        )
    return rows


# ── Summarize ──────────────────────────────────────────────────────────────────

def summarize(rows: List[Row], min_views: int = 0) -> Dict[str, Any]:
    """Compute summary stats, optionally filtering out low-view posts."""
    filtered = [row for row in rows if safe_num(row.get("reach_or_views")) >= min_views] if min_views > 0 else rows
    excluded = len(rows) - len(filtered)
    valid = [row for row in filtered if row["engagement_rate"] is not None]
    total_engagements = sum(safe_num(row["engagements"]) for row in filtered)
    total_denominator = sum(safe_num(row["reach_or_views"]) for row in filtered)
    return {
        "posts": len(rows),
        "excluded": excluded,
        "valid_posts": len(valid),
        "total_engagements": total_engagements,
        "total_denominator": total_denominator,
        "avg_er": mean([row["engagement_rate"] for row in valid]) if valid else None,
        "weighted_er": total_engagements / total_denominator if total_denominator else None,
    }


# ── Workbook builder ───────────────────────────────────────────────────────────

def build_workbook(
    output: Path,
    brand: str,
    blog_id: str,
    from_iso: str,
    to_iso: str,
    instagram_all: List[Row],
    tiktok: List[Row],
    ig_feed: List[Row],
    ig_reels: List[Row],
    youtube: List[Row],
    linkedin: List[Row],
    min_views: int = 0,
) -> Dict[str, Dict[str, Any]]:

    # Detect which networks are present
    has_tiktok   = bool(tiktok)
    has_youtube  = bool(youtube)
    has_linkedin = bool(linkedin)

    base_headers = [
        "Network",
        "Format",
        "Post ID",
        "Published At",
        "Timezone",
        "Type",
        "URL",
        "Caption",
        "Likes",
        "Comments",
        "Shares",
        "Saved / Clicks",
        "Total Engagements",
        "Denominator",
        "Reach / Views",
        "Engagement Rate",
        "Metricool ER",
        "Impressions",
        "Views",
        "Proof Formula",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_fill  = PatternFill("solid", fgColor="1F4E78")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    white_font   = Font(color="FFFFFF", bold=True)
    bold_font    = Font(bold=True)
    warning_font = Font(bold=True, color="7F6000")
    thin         = Side(style="thin", color="D9D9D9")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(row):
        for cell in row:
            cell.fill      = header_fill
            cell.font      = white_font
            cell.alignment = Alignment(horizontal="center")
            cell.border    = border

    def add_proof_sheet(title: str, rows: List[Row], apply_min_views: bool = False) -> None:
        sheet = wb.create_sheet(title)
        headers = base_headers + (["Min-Views Filter"] if apply_min_views else [])
        sheet.append(headers)
        style_header(sheet[1])
        for row in rows:
            denom       = safe_num(row.get("reach_or_views"))
            is_excluded = apply_min_views and min_views > 0 and denom <= min_views
            row_data = [
                row["network"],
                row["format"],
                row["post_id"],
                row["published_at"],
                row["timezone"],
                row["type"],
                row["url"],
                row["caption"],
                row["likes"],
                row["comments"],
                row["shares"],
                row["saved"],
                row["engagements"],
                row["denominator_label"],
                row["reach_or_views"],
                row["engagement_rate"],
                row["metricool_engagement_rate"],
                row["impressions"],
                row["views"],
                "=(Likes+Comments+Shares+Saved/Clicks)/(Reach or Views)",
            ]
            if apply_min_views:
                row_data.append(f"EXCLUDED (<= {min_views} views)" if is_excluded else "Included")
            sheet.append(clean_row(row_data))
        for xl_row in sheet.iter_rows(min_row=2):
            for cell in xl_row:
                cell.border    = border
                cell.alignment = Alignment(vertical="top", wrap_text=(cell.column in [7, 8, 14, 20]))
            # Grey out excluded rows
            if apply_min_views and min_views > 0 and xl_row[20].value and "EXCLUDED" in str(xl_row[20].value):
                excl_fill = PatternFill("solid", fgColor="F2F2F2")
                for cell in xl_row:
                    cell.fill = excl_fill
        for col in [16, 17]:
            for cells in sheet.iter_cols(min_col=col, max_col=col, min_row=2, max_row=sheet.max_row):
                for cell in cells:
                    if cell.value is not None:
                        cell.number_format = "0.00%"
        for col in list(range(9, 16)) + [18, 19]:
            for cells in sheet.iter_cols(min_col=col, max_col=col, min_row=2, max_row=sheet.max_row):
                for cell in cells:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0"
        widths = {
            1: 14, 2: 14, 3: 28, 4: 24, 5: 16, 6: 22, 7: 42, 8: 65,
            9: 10, 10: 10, 11: 10, 12: 12, 13: 18, 14: 44, 15: 15,
            16: 17, 17: 14, 18: 14, 19: 12, 20: 44,
        }
        if apply_min_views:
            widths[21] = 22
        for idx, width in widths.items():
            sheet.column_dimensions[get_column_letter(idx)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    # Add proof sheets
    add_proof_sheet("Instagram Proof", instagram_all)
    add_proof_sheet("TikTok Proof", tiktok, apply_min_views=True)
    add_proof_sheet("IG Feed Raw Proof", ig_feed)
    add_proof_sheet("IG Reels Raw Proof", ig_reels)
    if has_youtube:
        add_proof_sheet("YouTube Proof", youtube)
    if has_linkedin:
        add_proof_sheet("LinkedIn Proof", linkedin)

    # Summaries
    summaries: Dict[str, Dict[str, Any]] = {
        "Instagram Feed":   summarize(ig_feed,       0),
        "Instagram Reels":  summarize(ig_reels,      0),
        "Instagram Total":  summarize(instagram_all, 0),
        "TikTok":           summarize(tiktok,        min_views),
    }
    if has_youtube:
        summaries["YouTube"] = summarize(youtube, 0)
    if has_linkedin:
        summaries["LinkedIn"] = summarize(linkedin, 0)

    # ── Summary sheet content ──────────────────────────────────────────────────
    networks_label = "Instagram & TikTok"
    if has_youtube and has_linkedin:
        networks_label = "Instagram, TikTok, YouTube & LinkedIn"
    elif has_youtube:
        networks_label = "Instagram, TikTok & YouTube"
    elif has_linkedin:
        networks_label = "Instagram, TikTok & LinkedIn"

    ws["A1"] = f"{brand} — {networks_label} Engagement Rate Proof"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"] = "Brand"
    ws["B3"] = brand
    ws["A4"] = "Blog ID"
    ws["B4"] = blog_id
    ws["A5"] = "Source"
    ws["B5"] = "Metricool API v2 analytics post endpoints"
    ws["A6"] = "Date range"
    ws["B6"] = f"{from_iso} to {to_iso}"
    ws["A7"] = "Calculation"
    ws["B7"] = (
        "Engagement Rate = (Likes + Comments + Shares + Saved) / Reach. "
        "For TikTok and YouTube, Metricool exposes viewCount rather than reach; "
        "viewCount is used as the denominator/reach proxy. "
        "For LinkedIn, Impressions are used as the denominator; Clicks are included in engagements. "
        "Rows with denominator 0 are excluded from average ER."
    )
    ws["B7"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A8"] = "Min-Views Filter"
    ws["B8"] = (
        f"TikTok posts with {min_views} views/reach or fewer are excluded from ER averages "
        f"(shown in proof tabs as grey rows, marked EXCLUDED)."
        if min_views > 0 else "No minimum views filter applied (all posts included)."
    )
    ws["B8"].alignment = Alignment(wrap_text=True, vertical="top")

    for cell in ws["A3:A8"]:
        cell[0].font = bold_font

    ws.row_dimensions[7].height = 56
    ws.row_dimensions[8].height = 36

    # TikTok Saves warning box
    if has_tiktok:
        ws["A10"] = "⚠️  DATA LIMITATION"
        ws["A10"].font = warning_font
        ws["B10"] = TIKTOK_SAVES_NOTE
        ws["B10"].alignment = Alignment(wrap_text=True, vertical="top")
        ws["A10"].fill = warning_fill
        ws["B10"].fill = warning_fill
        ws.row_dimensions[10].height = 72
        ws["A10"].font = warning_font

    # Summary table
    sum_start = 13
    summary_headers = [
        "Network / Format",
        "Total Posts",
        "Posts with Denominator",
        "Total Engagements",
        "Total Reach / Views",
        "Average ER",
        "Weighted ER",
    ]
    for _ in range(sum_start - ws.max_row - 1):
        ws.append([])
    ws.append(summary_headers)
    style_header(ws[sum_start])

    for label, stats in summaries.items():
        ws.append(
            [
                label,
                stats["posts"],
                stats["valid_posts"],
                stats["total_engagements"],
                stats["total_denominator"],
                stats["avg_er"],
                stats["weighted_er"],
            ]
        )

    for row in ws.iter_rows(min_row=sum_start + 1, max_row=sum_start + len(summaries)):
        for cell in row:
            cell.border = border
        row[5].number_format = "0.00%"
        row[6].number_format = "0.00%"
        row[3].number_format = "#,##0"
        row[4].number_format = "#,##0"

    for col, width in {1: 24, 2: 14, 3: 24, 4: 20, 5: 20, 6: 14, 7: 14}.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A13"

    # Tabs legend
    tab_row = sum_start + len(summaries) + 2
    ws.cell(row=tab_row, column=1, value="Tabs").font = bold_font
    tab_desc = (
        "Instagram Proof combines feed + reels. "
        "IG Feed Raw Proof and IG Reels Raw Proof are included for auditability. "
        "TikTok Proof contains TikTok videos."
    )
    if has_youtube:
        tab_desc += " YouTube Proof contains YouTube videos."
    if has_linkedin:
        tab_desc += " LinkedIn Proof contains LinkedIn posts."
    ws.cell(row=tab_row, column=2, value=tab_desc).alignment = Alignment(wrap_text=True)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return summaries


# ── CLI ────────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build Metricool engagement-rate proof XLSX for Instagram, TikTok, "
            "YouTube, and/or LinkedIn."
        )
    )
    parser.add_argument("--blog-id", required=True, help="Metricool blogId/brand ID")
    parser.add_argument("--brand", help="Brand name. If omitted, fetched from Metricool.")
    parser.add_argument("--from", dest="from_date", required=True,
                        help="Start date: YYYY-MM-DD, ISO datetime, or 'today'")
    parser.add_argument("--to", dest="to_date", required=True,
                        help="End date: YYYY-MM-DD, ISO datetime, or 'today'")
    parser.add_argument(
        "--networks",
        default="instagram,tiktok",
        help=(
            "Comma-separated networks to fetch. "
            "Supported: instagram, tiktok, youtube, linkedin. "
            "Default: instagram,tiktok"
        ),
    )
    parser.add_argument("--output", help="Output XLSX path")
    parser.add_argument("--raw-dir", help="Directory for raw JSON responses")
    parser.add_argument("--no-raw", action="store_true", help="Do not save raw JSON responses")
    parser.add_argument("--no-instagram-reels", action="store_true",
                        help="Do not fetch Instagram reels")
    parser.add_argument(
        "--min-views",
        type=int,
        default=0,
        metavar="N",
        help=(
            "Exclude TikTok posts with N views/reach or fewer from ER averages. "
            "They still appear in proof tabs, greyed out and marked EXCLUDED. "
            "Recommended: 20. Default: 0 (no filter)."
        ),
    )
    parser.add_argument("--user-id", default=os.environ.get("METRICOOL_USER_ID", DEFAULT_USER_ID),
                        help="Metricool userId")
    return parser.parse_args()


def main() -> None:
    load_metricool_env()
    args = parse_args()
    token = os.environ.get("METRICOOL_TOKEN", "")
    if not token:
        raise SystemExit("METRICOOL_TOKEN is not set. Export METRICOOL_TOKEN or add it to a .env file.")

    from_iso   = normalize_date(args.from_date, end_of_day=False)
    to_iso     = normalize_date(args.to_date,   end_of_day=True)
    from_label = from_iso.split("T", 1)[0]
    to_label   = to_iso.split("T", 1)[0]

    client = MetricoolClient(args.blog_id, token, args.user_id)
    brand  = args.brand or client.brand_name() or f"Metricool Brand {args.blog_id}"

    networks = {part.strip().lower() for part in args.networks.split(",") if part.strip()}

    if args.output:
        output = Path(args.output)
    else:
        output = Path(
            f"{safe_filename(brand)}_Engagement_Rate_{from_label}_to_{to_label}.xlsx"
        )

    raw_dir: Optional[Path]
    if args.no_raw:
        raw_dir = None
    elif args.raw_dir:
        raw_dir = Path(args.raw_dir)
    else:
        raw_dir = Path(f"{safe_filename(brand)}_metricool_raw_{from_label}_to_{to_label}")

    ig_feed:  List[Row] = []
    ig_reels: List[Row] = []
    tiktok:   List[Row] = []
    youtube:  List[Row] = []
    linkedin: List[Row] = []

    if "instagram" in networks:
        print("Fetching Instagram feed posts...")
        feed_raw = client.fetch_paginated(
            "/v2/analytics/posts/instagram", from_iso, to_iso, raw_dir, "instagram_feed"
        )
        ig_feed = instagram_rows(feed_raw, "Feed Post")
        if not args.no_instagram_reels:
            print("Fetching Instagram reels...")
            reels_raw = client.fetch_paginated(
                "/v2/analytics/reels/instagram", from_iso, to_iso, raw_dir, "instagram_reels"
            )
            ig_reels = instagram_rows(reels_raw, "Reel")

    if "tiktok" in networks:
        print("Fetching TikTok posts...")
        print("  Note: TikTok Saves/Favorites not available via Metricool API — ER will exclude Saves.")
        tiktok_raw = client.fetch_paginated(
            "/v2/analytics/posts/tiktok", from_iso, to_iso, raw_dir, "tiktok"
        )
        tiktok = tiktok_rows(tiktok_raw)

    if "youtube" in networks:
        print("Fetching YouTube videos...")
        yt_raw = client.fetch_paginated(
            "/v2/analytics/posts/youtube", from_iso, to_iso, raw_dir, "youtube"
        )
        youtube = youtube_rows(yt_raw)

    if "linkedin" in networks:
        print("Fetching LinkedIn posts...")
        li_raw = client.fetch_paginated(
            "/v2/analytics/posts/linkedin", from_iso, to_iso, raw_dir, "linkedin"
        )
        linkedin = linkedin_rows(li_raw)

    def is_in_range(dt_str: str) -> bool:
        if not dt_str:
            return True
        return from_iso <= dt_str[:19] <= to_iso

    instagram_all = [r for r in ig_feed + ig_reels if is_in_range(str(r.get("published_at") or ""))]
    tiktok        = [r for r in tiktok if is_in_range(str(r.get("published_at") or ""))]
    youtube       = [r for r in youtube if is_in_range(str(r.get("published_at") or ""))]
    linkedin      = [r for r in linkedin if is_in_range(str(r.get("published_at") or ""))]

    instagram_all = sorted(instagram_all, key=lambda r: r.get("published_at") or "", reverse=True)
    tiktok        = sorted(tiktok,        key=lambda r: r.get("published_at") or "", reverse=True)
    youtube       = sorted(youtube,       key=lambda r: r.get("published_at") or "", reverse=True)
    linkedin      = sorted(linkedin,      key=lambda r: r.get("published_at") or "", reverse=True)

    if args.min_views > 0:
        print(f"Min-views filter applied: posts with < {args.min_views} views excluded from ER averages.")

    summaries = build_workbook(
        output, brand, args.blog_id, from_iso, to_iso,
        instagram_all, tiktok, ig_feed, ig_reels, youtube, linkedin,
        min_views=args.min_views,
    )

    print(f"\nSaved: {output}")
    if raw_dir:
        print(f"Raw JSON: {raw_dir}")
    print()
    for label, stats in summaries.items():
        avg      = stats["avg_er"]
        weighted = stats["weighted_er"]
        excl     = stats.get("excluded", 0)
        avg_s    = f"{avg:.2%}" if avg is not None else "n/a"
        w_s      = f"{weighted:.2%}" if weighted is not None else "n/a"
        excl_s   = f", excluded={excl}" if args.min_views > 0 else ""
        print(
            f"{label}: posts={stats['posts']}{excl_s}, "
            f"valid={stats['valid_posts']}, avg_er={avg_s}, weighted_er={w_s}"
        )


if __name__ == "__main__":
    main()
