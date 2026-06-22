#!/usr/bin/env python3
"""
Build an Instagram engagement-rate proof workbook from Supabase competitor_posts data.

Usage:
    python scripts/export_ig_xlsx.py \\
        --profile-id 1446c72e-... \\
        --brand "Anmum" \\
        --from 2026-05-01 \\
        --to 2026-05-31 \\
        --output anmum_may2026.xlsx

Requires: pip install openpyxl
"""
import argparse
import sys
from pathlib import Path
from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Ensure we can import ig_utils from the scripts directory
_script_dir = Path(__file__).resolve().parent
sys.path.insert(0, str(_script_dir))
from ig_utils import get_supabase, DEFAULT_SUPABASE_URL, DEFAULT_SUPABASE_KEY, resolve_output_path

# Styling
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
ALT_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
BORDER = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7'),
)
TITLE_FONT = Font(name='Calibri', bold=True, size=14)
SUBTITLE_FONT = Font(name='Calibri', bold=True, size=11)


def style_header_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER


def style_data_cell(cell, alt=False):
    cell.font = Font(name='Calibri', size=10)
    cell.border = BORDER
    cell.alignment = Alignment(vertical='center', wrap_text=True)
    if alt:
        cell.fill = ALT_FILL


def build_xlsx(posts, brand, from_date, to_date):
    """Build workbook with Summary and Proof sheets."""
    wb = Workbook()

    # ── Summary Sheet ──
    ws_sum = wb.active
    ws_sum.title = 'Summary'
    ws_sum.sheet_properties.tabColor = '2F5496'

    ws_sum.cell(row=1, column=1, value=f'{brand} — Instagram Engagement Report').font = TITLE_FONT
    ws_sum.cell(row=2, column=1, value=f'{from_date} to {to_date}').font = SUBTITLE_FONT
    ws_sum.merge_cells('A1:D1')
    ws_sum.merge_cells('A2:D2')

    # Stats
    total_posts = len(posts)
    total_likes = sum(p['raw'].get('likes', 0) or 0 for p in posts)
    total_comments = sum(p['raw'].get('comments', 0) or 0 for p in posts)
    total_engagements = total_likes + total_comments
    collab_count = sum(1 for p in posts if p['raw'].get('is_collab', False))
    video_count = sum(1 for p in posts if p['raw'].get('media_type') == 2)
    image_count = sum(1 for p in posts if p['raw'].get('media_type') == 1)
    carousel_count = sum(1 for p in posts if p['raw'].get('media_type') == 8)

    # Engagement rate (per post, then average)
    ers = []
    for p in posts:
        likes = p['raw'].get('likes', 0) or 0
        comments = p['raw'].get('comments', 0) or 0
        views = p['raw'].get('video_view_count')
        denom = views or (likes + comments if (likes + comments) > 0 else 1)
        if denom and denom > 0:
            ers.append((likes + comments) / denom * 100)

    avg_er = sum(ers) / len(ers) if ers else 0
    # Weighted ER = total engagements / total reach proxy (using views as proxy)
    total_views = sum((p['raw'].get('video_view_count') or 0) for p in posts if p['raw'].get('video_view_count'))
    if total_views > 0:
        weighted_er = total_engagements / total_views * 100
    else:
        total_denom = sum(
            (p['raw'].get('likes', 0) or 0) + (p['raw'].get('comments', 0) or 0)
            for p in posts
        )
        weighted_er = (total_engagements / total_denom * 100) if total_denom > 0 else 0

    stats = [
        ('Total Posts', total_posts),
        ('Date Range', f'{from_date} to {to_date}'),
        ('Total Likes', total_likes),
        ('Total Comments', total_comments),
        ('Total Engagements', total_engagements),
        ('Average ER%', f'{avg_er:.2f}%'),
        ('Weighted ER%', f'{weighted_er:.2f}%'),
        ('Collab Posts', collab_count),
        ('Videos', video_count),
        ('Images', image_count),
        ('Carousels', carousel_count),
    ]

    for i, (label, val) in enumerate(stats, start=4):
        ws_sum.cell(row=i, column=1, value=label).font = SUBTITLE_FONT
        ws_sum.cell(row=i, column=2, value=val).font = Font(name='Calibri', size=11)
        ws_sum.cell(row=i, column=1).border = BORDER
        ws_sum.cell(row=i, column=2).border = BORDER

    ws_sum.column_dimensions['A'].width = 20
    ws_sum.column_dimensions['B'].width = 25

    # ── Proof Sheet ──
    ws = wb.create_sheet('Instagram Proof')
    ws.sheet_properties.tabColor = '70AD47'

    headers = [
        'Date', 'Post Type', 'Shortcode', 'URL', 'Likes', 'Comments',
        'Video Views', 'Engagements', 'ER%', 'Collab?', 'Tagged Users',
        'Caption Preview',
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    for i, p in enumerate(posts, start=2):
        raw = p.get('raw', {})
        likes = raw.get('likes', 0) or 0
        comments = raw.get('comments', 0) or 0
        views = raw.get('video_view_count')
        engagements = likes + comments
        denom = views or engagements if engagements > 0 else 1
        er = (engagements / denom * 100) if denom > 0 else 0

        tagged = raw.get('tagged_users', []) or []
        cap = (raw.get('caption') or '')[:80].replace('\n', ' ')

        row_data = [
            p.get('created_at', '')[:10],
            raw.get('post_type', '?'),
            raw.get('code', ''),
            raw.get('post_url', ''),
            likes,
            comments,
            views or '',
            engagements,
            f'{er:.1f}%',
            'Yes' if raw.get('is_collab') else '',
            ', '.join(tagged) if tagged else '',
            cap,
        ]

        alt = (i % 2 == 0)
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            style_data_cell(cell, alt=alt)

    # Column widths
    widths = [12, 10, 14, 38, 8, 9, 11, 12, 8, 8, 20, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return wb


def main():
    parser = argparse.ArgumentParser(description='Export IG engagement report from Supabase data')
    parser.add_argument('--profile-id', required=True, help='competitor_profiles UUID')
    parser.add_argument('--brand', required=True, help='Brand name for report title')
    parser.add_argument('--from', dest='from_date', required=True)
    parser.add_argument('--to', dest='to_date', required=True)
    parser.add_argument('--output', default=None, help='Output .xlsx path')
    parser.add_argument('--supabase-url', default=DEFAULT_SUPABASE_URL)
    parser.add_argument('--supabase-key', default=DEFAULT_SUPABASE_KEY)
    args = parser.parse_args()

    print(f'Fetching posts for {args.brand} ({args.from_date} to {args.to_date})...')
    supabase = get_supabase(args.supabase_url, args.supabase_key)
    posts = get_supabase().table('competitor_posts') \
        .select('*') \
        .eq('profile_id', args.profile_id) \
        .eq('platform', 'instagram') \
        .gte('created_at', f'{args.from_date}T00:00:00') \
        .lte('created_at', f'{args.to_date}T23:59:59') \
        .order('created_at', desc=True) \
        .execute()

    posts = posts.data or []
    print(f'  Found {len(posts)} posts')

    if not posts:
        print('No posts to export. Exiting.')
        return

    wb = build_xlsx(posts, args.brand, args.from_date, args.to_date)

    default_name = f'{args.brand.lower().replace(" ", "_")}_ig_{args.from_date}_{args.to_date}.xlsx'
    output = resolve_output_path(args.output or default_name)
    wb.save(output)
    print(f'  Saved: {output}')

    # Quick summary
    likes = sum(p['raw'].get('likes', 0) or 0 for p in posts)
    comments = sum(p['raw'].get('comments', 0) or 0 for p in posts)
    collab = sum(1 for p in posts if p['raw'].get('is_collab', False))
    print(f'\nSummary: {len(posts)} posts, {likes} likes, {comments} comments, {collab} collab posts')


if __name__ == '__main__':
    main()
